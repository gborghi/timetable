"""Dashboard endpoints: dataset state already lives in routers/dataset.py;
this router hosts the `/api/dashboard/graph` endpoint that powers the
Network panel, plus Import/Export DB and snapshot management."""
from __future__ import annotations

import csv
import datetime as dt
import hashlib
import io
import json
import os
import re
import shutil
import zipfile
from typing import Any

from fastapi import (APIRouter, Depends, File, HTTPException, Query,
                      UploadFile)
from fastapi.responses import JSONResponse, StreamingResponse
from sqlalchemy import inspect, text
from sqlalchemy.orm import Session

from .. import db as db_mod
from ..db import get_db
from ..services.graph import build_graph
from ..utils.ttl_cache import cached as ttl_cached

router = APIRouter(prefix="/api/dashboard", tags=["dashboard"])


# ----------------------------- helpers --------------------------------

# webui/data/snapshots/
DATA_DIR = db_mod.DATA_DIR
SNAPSHOTS_DIR = os.path.join(DATA_DIR, "snapshots")
os.makedirs(SNAPSHOTS_DIR, exist_ok=True)


def _safe_filename(name: str) -> str:
    """Strip any path component and reject anything outside [A-Za-z0-9._-]."""
    base = os.path.basename(name or "")
    if not base:
        raise HTTPException(400, "filename vuoto")
    if re.search(r"[^A-Za-z0-9._-]", base):
        raise HTTPException(400, "filename non valido")
    return base


def _alembic_revision() -> str | None:
    try:
        with db_mod.engine.connect() as conn:
            row = conn.execute(
                text("SELECT version_num FROM alembic_version")
            ).first()
            return str(row[0]) if row else None
    except Exception:
        return None


def _table_csv(conn, table: str) -> bytes:
    """Dump one table as CSV bytes using the live engine connection."""
    res = conn.execute(text(f'SELECT * FROM "{table}"'))
    cols = list(res.keys())
    buf = io.StringIO()
    w = csv.writer(buf, lineterminator="\r\n")
    w.writerow(cols)
    for row in res:
        w.writerow([("" if v is None else v) for v in row])
    return b"\xef\xbb\xbf" + buf.getvalue().encode("utf-8")


def _build_export_zip(*, schema_only: bool = False) -> tuple[bytes, dict]:
    """Build the export zip in memory. Returns (zip_bytes, metadata_dict).

    Layout inside the zip:
      database.db                        -- raw SQLite file (when SQLite)
      metadata.json                      -- schema_version + counts + hashes
      tables/<table>.csv                 -- one CSV per table
                                            (skipped when schema_only=True)
    """
    insp = inspect(db_mod.engine)
    tables = sorted(insp.get_table_names())
    row_counts: dict[str, int] = {}
    csv_hashes: dict[str, str] = {}

    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as z:
        # Raw SQLite file (when applicable). Fast atomic restore path.
        if db_mod.IS_SQLITE and db_mod.DB_PATH and os.path.exists(db_mod.DB_PATH):
            # WAL mode keeps recently-committed pages in `.db-wal` until
            # the next checkpoint. Force a TRUNCATE checkpoint so every
            # committed page is folded into the main .db file before we
            # snapshot it; otherwise the export could miss the latest
            # writes.
            try:
                with db_mod.engine.connect() as conn:
                    conn.exec_driver_sql("PRAGMA wal_checkpoint(TRUNCATE)")
            except Exception:
                # Non-fatal: a missing/older WAL just means there's
                # nothing to checkpoint. Carry on with the snapshot.
                pass
            with open(db_mod.DB_PATH, "rb") as f:
                z.writestr("database.db", f.read())
        # Per-table CSVs (always for documentation/portability)
        with db_mod.engine.connect() as conn:
            for t in tables:
                try:
                    cnt = conn.execute(
                        text(f'SELECT COUNT(*) FROM "{t}"')
                    ).scalar() or 0
                except Exception:
                    cnt = 0
                row_counts[t] = int(cnt)
                if schema_only:
                    continue
                try:
                    body = _table_csv(conn, t)
                except Exception:
                    continue
                z.writestr(f"tables/{t}.csv", body)
                csv_hashes[t] = hashlib.sha256(body).hexdigest()
        meta = {
            "kind": "pitantum-db-export",
            "schema_version": _alembic_revision(),
            "schema_only": bool(schema_only),
            "exported_at": dt.datetime.now().isoformat(timespec="seconds"),
            "tables": tables,
            "row_counts": row_counts,
            "csv_sha256": csv_hashes,
            "is_sqlite": db_mod.IS_SQLITE,
        }
        z.writestr("metadata.json", json.dumps(meta, indent=2))
    return buf.getvalue(), meta


def _restore_zip(blob: bytes) -> dict:
    """Apply an export zip to the running DB. Strategy:
       - if the zip contains `database.db` AND we're on SQLite,
         atomically replace the live DB file (closes connections,
         copies, reconnects).
       - otherwise: not yet supported (would require row-by-row import
         from CSV). Returns 400.
    """
    try:
        z = zipfile.ZipFile(io.BytesIO(blob))
    except zipfile.BadZipFile:
        raise HTTPException(400, "il file caricato non e' un zip valido")
    names = z.namelist()
    if "metadata.json" not in names:
        raise HTTPException(
            400,
            "metadata.json mancante: il file non sembra un export piTantum.",
        )
    try:
        meta = json.loads(z.read("metadata.json").decode("utf-8"))
    except Exception as e:
        raise HTTPException(400, f"metadata.json non parseable: {e}")
    if meta.get("kind") != "pitantum-db-export":
        raise HTTPException(
            400, f"export non riconosciuto: kind={meta.get('kind')}"
        )
    if not db_mod.IS_SQLITE:
        raise HTTPException(
            400,
            "import-db: solo SQLite supportato (per Postgres serve un "
            "pg_restore lato server, non disponibile da qui).",
        )
    if "database.db" not in names:
        raise HTTPException(
            400,
            "database.db assente nel zip; il restore CSV-per-CSV non e' "
            "ancora supportato.",
        )
    new_db = z.read("database.db")
    # Atomically swap the SQLite file
    db_path = db_mod.DB_PATH
    if not db_path:
        raise HTTPException(500, "DB path non risolvibile")
    # Close pooled connections so Windows lets us replace the file
    db_mod.engine.dispose()
    backup_path = db_path + ".pre_import_backup"
    try:
        if os.path.exists(db_path):
            shutil.copy2(db_path, backup_path)
        # WAL mode side effect: SQLite leaves `.db-wal` and `.db-shm`
        # sidecar files next to the main .db. After a successful
        # dispose() the sidecars belong to the OLD database; if we
        # leave them on disk SQLite will re-apply their pages on top
        # of the freshly written .db at the next open and corrupt
        # the import. Remove them before swapping.
        for suffix in ("-wal", "-shm"):
            sidecar = db_path + suffix
            if os.path.exists(sidecar):
                try:
                    os.remove(sidecar)
                except OSError:
                    pass
        with open(db_path, "wb") as f:
            f.write(new_db)
    except Exception as e:
        # Best-effort rollback
        if os.path.exists(backup_path):
            shutil.copy2(backup_path, db_path)
        raise HTTPException(500, f"errore swap DB: {e}")
    finally:
        # New connections will auto-open against the new file
        pass
    # Bump mutation epoch so caches invalidate
    try:
        from ..utils.ttl_cache import bump_mutation
        bump_mutation()
    except Exception:
        pass
    return {
        "ok": True,
        "schema_version": meta.get("schema_version"),
        "exported_at": meta.get("exported_at"),
        "tables": meta.get("tables"),
        "row_counts": meta.get("row_counts"),
        "backup": backup_path,
    }


# ----------------------------- endpoints ------------------------------


@router.get("/graph")
def get_graph(
    mode: str = Query(
        "classes",
        description=(
            "'classes' -> nodes are SchoolClass, edges weighted by shared "
            "teachers; 'teachers' -> nodes are Teacher, edges weighted by "
            "shared classes."
        ),
        pattern="^(classes|teachers)$",
    ),
    db: Session = Depends(get_db),
):
    """Network graph. Server-side cache 60s + mutation-aware; plus
    Cache-Control: no-store on response so the browser never serves a
    stale frame after an import."""
    try:
        body = ttl_cached(
            f"dashboard.graph.{mode}",
            ttl_s=60.0,
            mutation_aware=True,
            compute=lambda: build_graph(db, mode),
        )
        return JSONResponse(
            content=body,
            headers={"Cache-Control": "no-store, max-age=0"},
        )
    except ValueError as e:
        raise HTTPException(400, str(e))


# ----- Import / Export DB ----------------------------------------------

@router.get("/export-db")
def export_db(
    schema_only: bool = Query(
        False,
        description=(
            "If true, omits the table data — useful as a 'template' "
            "starter package for a different school."
        ),
    ),
):
    """Build a zip containing:
      - database.db    (raw SQLite, when applicable)
      - tables/*.csv   (one CSV per table, BOM+UTF-8)
      - metadata.json  (schema_version + row counts + sha256)
    """
    body, meta = _build_export_zip(schema_only=schema_only)
    ts = dt.datetime.now().strftime("%Y%m%d_%H%M%S")
    fn = f"pitantum_export_{ts}.zip"
    return StreamingResponse(
        io.BytesIO(body),
        media_type="application/zip",
        headers={
            "Content-Disposition": f'attachment; filename="{fn}"',
            "X-Pitantum-Schema-Version": meta.get("schema_version") or "",
        },
    )


@router.post("/import-db")
async def import_db(file: UploadFile = File(...)):
    """Upload a zip previously produced by /api/dashboard/export-db
    and replace the running DB with it. SQLite-only path: writes a
    .pre_import_backup file next to the live DB before the swap.
    """
    blob = await file.read()
    if not blob:
        raise HTTPException(400, "file vuoto")
    return _restore_zip(blob)


@router.post("/snapshot/create")
def snapshot_create():
    """Save a timestamped snapshot zip in webui/data/snapshots/ and
    return the filename + size."""
    body, meta = _build_export_zip(schema_only=False)
    ts = dt.datetime.now().strftime("%Y%m%d_%H%M%S")
    fn = f"snapshot_{ts}.zip"
    path = os.path.join(SNAPSHOTS_DIR, fn)
    with open(path, "wb") as f:
        f.write(body)
    return {
        "ok": True,
        "filename": fn,
        "path": path,
        "size_bytes": len(body),
        "schema_version": meta.get("schema_version"),
    }


@router.get("/snapshot/list")
def snapshot_list():
    """List all snapshot zips in webui/data/snapshots/."""
    out: list[dict[str, Any]] = []
    if not os.path.isdir(SNAPSHOTS_DIR):
        return out
    for fn in sorted(os.listdir(SNAPSHOTS_DIR), reverse=True):
        if not fn.lower().endswith(".zip"):
            continue
        full = os.path.join(SNAPSHOTS_DIR, fn)
        try:
            stat = os.stat(full)
        except OSError:
            continue
        out.append({
            "filename": fn,
            "size_bytes": stat.st_size,
            "modified_at": dt.datetime.fromtimestamp(
                stat.st_mtime
            ).isoformat(timespec="seconds"),
        })
    return out


@router.post("/snapshot/restore/{filename}")
def snapshot_restore(filename: str):
    fn = _safe_filename(filename)
    full = os.path.join(SNAPSHOTS_DIR, fn)
    if not os.path.isfile(full):
        raise HTTPException(404, f"snapshot {fn} non trovato")
    with open(full, "rb") as f:
        return _restore_zip(f.read())


@router.delete("/snapshot/{filename}")
def snapshot_delete(filename: str):
    fn = _safe_filename(filename)
    full = os.path.join(SNAPSHOTS_DIR, fn)
    if not os.path.isfile(full):
        raise HTTPException(404, f"snapshot {fn} non trovato")
    os.remove(full)
    return {"ok": True, "deleted": fn}


# =====================================================================
# Constraints import / export
# =====================================================================
#
# The Dashboard "Vincoli" sub-card exposes:
#   POST /api/dashboard/constraints/import-stress
#       body: {profile: small|medium|big|huge|superhuge|mega}
#       reads experiments/stress_constraints/<profile>/*.json
#       and bulk-creates the constraints in the live DB.
#   POST /api/dashboard/constraints/import-file
#       multipart upload (JSON or xlsx); validates each record and
#       imports the valid ones; returns a per-record report.
#   GET  /api/dashboard/constraints/export?format=json|xlsx
#       returns a file with the constraint records currently in the
#       DB, in the same shape the import endpoints accept.
#   DELETE /api/dashboard/constraints/all?confirm=true
#       wipes every constraint table; returns counts.
# ---------------------------------------------------------------------

from .. import models as _models   # noqa: E402

_STRESS_PROFILES = ("small", "medium", "big", "huge", "superhuge", "mega")


def _stress_dir(profile: str) -> str:
    """Resolve the stress dataset folder for a profile. Falls back to
    walking up from the backend dir to find experiments/."""
    profile = (profile or "").lower().strip()
    if profile not in _STRESS_PROFILES:
        raise HTTPException(
            400, f"profilo {profile!r} non valido; "
                  f"usa uno tra {list(_STRESS_PROFILES)}.")
    here = os.path.dirname(os.path.abspath(__file__))
    # walk up to find experiments/
    cur = here
    for _ in range(8):
        cand = os.path.join(cur, "experiments", "stress_constraints",
                            profile)
        if os.path.isdir(cand):
            return cand
        cur = os.path.dirname(cur)
    raise HTTPException(
        500,
        f"cartella stress_constraints/{profile} non trovata; "
        "probabile esecuzione in un working dir non standard.")


def _resolve_owner(db: Session, scope: str, pattern: str) -> int | None:
    """Resolve `owner_pattern` for stress datasets to an entity id in
    the live DB. Mirrors experiments/load_stress_constraints.py."""
    if not pattern:
        return None
    pattern = pattern.strip()
    nth_word_to_idx = {n: i for i, n in enumerate(
        ["first", "second", "third", "fourth", "fifth", "sixth",
         "seventh", "eighth", "ninth", "tenth", "eleventh",
         "twelfth"])}
    if scope == "teacher":
        rows = (db.query(_models.Teacher)
                  .order_by(_models.Teacher.id.asc()).all())
        if pattern.endswith("_teacher"):
            tok = pattern[:-len("_teacher")]
            idx = nth_word_to_idx.get(tok)
            if idx is not None and idx < len(rows):
                return rows[idx].id
        for t in rows:
            for cand in (getattr(t, "name", None),
                          getattr(t, "nickname", None),
                          getattr(t, "display_name", None)):
                if cand and cand == pattern:
                    return t.id
            ln = getattr(t, "last_name", "") or ""
            fn = getattr(t, "first_name", "") or ""
            if pattern == f"{ln} {fn}".strip():
                return t.id
            if pattern == f"{ln} {fn[:1]}.".strip(" ."):
                return t.id
        return None
    if scope == "classroom":
        rooms = (db.query(_models.Classroom)
                   .order_by(_models.Classroom.id.asc()).all())
        labs = sorted([r for r in rooms
                       if (getattr(r, "kind", "") or "").lower()
                          .startswith("lab")],
                      key=lambda r: r.id)
        nth_lab = {"first_lab": 0, "second_lab": 1, "third_lab": 2,
                   "fourth_lab": 3}
        if pattern in nth_lab and nth_lab[pattern] < len(labs):
            return labs[nth_lab[pattern]].id
        if pattern == "gym":
            for r in rooms:
                k = (getattr(r, "kind", "") or "").lower()
                n = (getattr(r, "name", "") or "").lower()
                if "palest" in k or "palest" in n:
                    return r.id
        if pattern == "main_room":
            for r in rooms:
                n = (getattr(r, "name", "") or "").lower()
                k = (getattr(r, "kind", "") or "").lower()
                if "magna" in n or k == "aula_magna":
                    return r.id
        for r in rooms:
            if pattern == getattr(r, "name", None):
                return r.id
        return None
    if scope == "class":
        classes = (db.query(_models.SchoolClass)
                     .order_by(_models.SchoolClass.id.asc()).all())
        if "_class_year_" in pattern:
            tok, _, year_str = pattern.partition("_class_year_")
            try:
                year = int(year_str)
            except ValueError:
                return None
            idx = nth_word_to_idx.get(tok, 0)
            in_year = sorted([c for c in classes
                              if getattr(c, "year", None) == year],
                             key=lambda c: c.id)
            if idx < len(in_year):
                return in_year[idx].id
            return None
        for c in classes:
            if pattern == getattr(c, "name", None):
                return c.id
        return None
    if scope == "curriculum":
        for c in db.query(_models.Curriculum).all():
            if pattern == getattr(c, "name", None):
                return c.id
        return None
    return None


_DAY_TOKENS = ("lun", "mar", "mer", "gio", "ven", "sab")
# Pre-compile a regex to expand bare day tokens (used as a "whole day"
# shortcut in the stress datasets) into the explicit all-hours
# disjunction the DSL grammar requires. Match a day word that is NOT
# followed by a digit (i.e. NOT 'lun8' or 'mar11'). Word boundaries
# guard against partial matches in 'martedi'.
import re as _re   # local alias to keep top-of-file imports clean
_BARE_DAY_RE = _re.compile(
    r"(?<![A-Za-z'])(lun|mar|mer|gio|ven|sab)(?![A-Za-z'\d])",
    flags=_re.IGNORECASE,
)


def _expand_bare_days(expr: str) -> str:
    """Expand 'mai sab' / 'lun OR mar' shortcut into the explicit
    all-hours disjunction the DSL parser understands. Idempotent: a
    token already followed by a digit (e.g. 'sab8') is untouched.
    """
    if not expr:
        return expr

    def _sub(m):
        d = m.group(1).lower()
        return ("(" + " OR ".join(f"{d}{h}" for h in (8, 9, 10, 11, 12, 13))
                + ")")
    return _BARE_DAY_RE.sub(_sub, expr)


def _create_constraint_via_dispatcher(db: Session, body: dict) -> dict:
    """Bridge function: forward a constraint creation payload through
    the same logic as POST /api/constraints. Returns
    {ok, kind, id, scope, error}."""
    from . import constraints as _crouter
    from ..schemas import ConstraintCreateIn
    try:
        payload = ConstraintCreateIn(**body)
        # Reuse the dispatcher; it commits internally.
        out = _crouter.create_constraint(payload, db=db)
        # The dispatcher returns ConstraintCreateOut OR a dict-ish
        # equivalent depending on FastAPI mode; normalize.
        if hasattr(out, "model_dump"):
            return {**out.model_dump(), "ok": True}
        if isinstance(out, dict):
            return {**out, "ok": True}
        return {"ok": True, "kind": body.get("kind"),
                "id": getattr(out, "id", None)}
    except HTTPException as e:
        return {"ok": False, "error": f"HTTP {e.status_code}: {e.detail}"}
    except Exception as e:
        return {"ok": False, "error": str(e)}


# ----- Import: stress profile -----------------------------------------

from pydantic import BaseModel as _BM   # noqa: E402


class ImportStressIn(_BM):
    profile: str


@router.post("/constraints/import-stress")
def constraints_import_stress(payload: ImportStressIn,
                                db: Session = Depends(get_db)):
    """Load a curated stress profile from
    `experiments/stress_constraints/<profile>/` into the live DB.

    Returns a per-record report:
      {profile, n_loaded, n_errors, by_category, intentional_conflicts,
       results: [{dataset_id, ok, db_id?, error?}, ...]}
    """
    src_dir = _stress_dir(payload.profile)
    files = ["teacher_constraints.json",
             "classroom_constraints.json",
             "relational_constraints.json"]
    dataset = []
    for fn in files:
        full = os.path.join(src_dir, fn)
        if not os.path.exists(full):
            continue
        try:
            data = json.load(open(full, encoding="utf-8"))
        except Exception as e:
            raise HTTPException(500, f"errore parsing {fn}: {e}")
        for c in data.get("constraints", []):
            dataset.append(c)
    if not dataset:
        raise HTTPException(404, f"nessun record per il profilo "
                                 f"{payload.profile!r}")
    by_category = {"teacher": 0, "classroom": 0, "relational": 0}
    n_intentional = 0
    n_ok = 0
    n_err = 0
    results = []
    for rec in dataset:
        scope = (rec.get("scope") or "").lower()
        # Tally category from id prefix (e.g. small_t_001 -> teacher)
        rid = rec.get("id") or ""
        if "_t_" in rid:
            by_category["teacher"] += 1
        elif "_r_" in rid:
            by_category["classroom"] += 1
        elif "_x_" in rid:
            by_category["relational"] += 1
        if rec.get("intentionally_conflicting"):
            n_intentional += 1
        owner_id = _resolve_owner(db, scope, rec.get("owner_pattern"))
        if owner_id is None:
            n_err += 1
            results.append({"dataset_id": rid, "ok": False,
                            "error": (f"owner_pattern "
                                      f"'{rec.get('owner_pattern')}'"
                                      f" non risolvibile (scope={scope})")})
            continue
        body = {
            "scope": scope,
            "kind": "logical",
            "level": rec.get("level"),
            "owner_id": owner_id,
            "expression": _expand_bare_days(rec.get("expression") or ""),
        }
        if rec.get("weight") is not None:
            body["weight"] = rec["weight"]
        out = _create_constraint_via_dispatcher(db, body)
        if out.get("ok"):
            n_ok += 1
            results.append({"dataset_id": rid, "ok": True,
                            "db_id": out.get("id")})
        else:
            n_err += 1
            results.append({"dataset_id": rid, "ok": False,
                            "error": out.get("error")})
    db.commit()
    try:
        from ..utils.ttl_cache import bump_mutation
        bump_mutation()
    except Exception:
        pass
    return {
        "ok": True,
        "profile": payload.profile,
        "n_total": len(dataset),
        "n_loaded": n_ok,
        "n_errors": n_err,
        "by_category": by_category,
        "intentional_conflicts": n_intentional,
        "results": results,
    }


# ----- Import: custom file (JSON / xlsx) ------------------------------

def _normalize_record(rec: dict) -> dict | None:
    """Convert one record from the canonical JSON-import shape to a
    `ConstraintCreateIn` body dict.

    Canonical shape (per spec):
      {kind, scope, owner_name, level, expression, soft_penalty?,
       description?}

    Returns None if the record is missing required fields. The
    `owner_name` -> `owner_id` resolution happens later (we need a
    DB session)."""
    if not isinstance(rec, dict):
        return None
    scope = (rec.get("scope") or "").strip().lower()
    kind = (rec.get("kind") or "").strip().lower()
    if not scope or not kind:
        return None
    body = {
        "scope": scope,
        "kind": kind,
        "level": (rec.get("level") or "hard").strip().lower(),
    }
    if rec.get("expression") is not None:
        body["expression"] = _expand_bare_days(str(rec["expression"]))
    if rec.get("soft_penalty") is not None:
        body["weight"] = int(rec["soft_penalty"])
    elif rec.get("weight") is not None:
        body["weight"] = int(rec["weight"])
    if rec.get("description"):
        body["label"] = str(rec["description"])
    if rec.get("subject"):
        body["subject"] = str(rec["subject"])
    if rec.get("day") is not None:
        body["day"] = int(rec["day"])
    if rec.get("hour") is not None:
        body["hour"] = int(rec["hour"])
    if rec.get("year_filter") is not None:
        body["year_filter"] = int(rec["year_filter"])
    body["_owner_name"] = (rec.get("owner_name") or "").strip()
    return body


def _resolve_owner_by_name(db: Session, scope: str,
                            owner_name: str) -> int | None:
    """Lenient name-based resolution for import-file (no nth_*
    sugar — that's stress-profile specific)."""
    if not owner_name:
        return None
    q = owner_name.strip()
    if scope == "teacher":
        for t in db.query(_models.Teacher).all():
            for cand in (getattr(t, "name", None),
                          getattr(t, "nickname", None),
                          getattr(t, "display_name", None)):
                if cand and cand == q:
                    return t.id
            ln = getattr(t, "last_name", "") or ""
            fn = getattr(t, "first_name", "") or ""
            for combo in (
                f"{ln} {fn}", f"{fn} {ln}",
                f"{ln} {fn[:1]}.", f"{fn[:1]}. {ln}",
            ):
                if combo.strip(" .") == q.strip(" ."):
                    return t.id
        return None
    if scope == "class":
        for c in db.query(_models.SchoolClass).all():
            if getattr(c, "name", None) == q:
                return c.id
        return None
    if scope == "classroom":
        for r in db.query(_models.Classroom).all():
            if getattr(r, "name", None) == q:
                return r.id
        return None
    if scope == "curriculum":
        for c in db.query(_models.Curriculum).all():
            if getattr(c, "name", None) == q:
                return c.id
        return None
    return None


def _parse_import_blob(blob: bytes, content_type: str | None,
                        filename: str | None) -> list[dict]:
    """Decode a .json or .xlsx upload to a list of canonical record
    dicts. Raises HTTPException on parse error."""
    name = (filename or "").lower()
    is_json = (name.endswith(".json")
                or (content_type or "").startswith("application/json")
                or (name == "" and blob.lstrip().startswith(b"[")))
    is_xlsx = (name.endswith(".xlsx")
                or (content_type or "").endswith("spreadsheetml.sheet"))
    if is_json:
        try:
            data = json.loads(blob.decode("utf-8"))
        except Exception as e:
            raise HTTPException(400, f"JSON non valido: {e}")
        if isinstance(data, dict) and "constraints" in data:
            data = data["constraints"]
        if not isinstance(data, list):
            raise HTTPException(400,
                "JSON deve essere una lista di record.")
        return data
    if is_xlsx:
        try:
            from openpyxl import load_workbook
        except Exception:
            raise HTTPException(
                500,
                "openpyxl non installato; impossibile leggere xlsx.")
        try:
            wb = load_workbook(filename=io.BytesIO(blob),
                                data_only=True, read_only=True)
        except Exception as e:
            raise HTTPException(400, f"xlsx non valido: {e}")
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return []
        headers = [str(h or "").strip() for h in rows[0]]
        out = []
        for row in rows[1:]:
            if all(c is None or c == "" for c in row):
                continue
            rec = {}
            for h, v in zip(headers, row):
                if not h:
                    continue
                rec[h] = v
            out.append(rec)
        return out
    raise HTTPException(
        400, "formato non riconosciuto: usa .json o .xlsx.")


@router.post("/constraints/import-file")
async def constraints_import_file(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
):
    """Upload a JSON or xlsx file with constraint records and import the
    valid ones. Returns:
      {n_total, n_loaded, n_errors, errors: [{idx, error, record}, ...]}
    """
    blob = await file.read()
    if not blob:
        raise HTTPException(400, "file vuoto")
    raw_records = _parse_import_blob(blob, file.content_type,
                                       file.filename)
    n_loaded = 0
    n_errors = 0
    errors = []
    for idx, rec in enumerate(raw_records):
        body = _normalize_record(rec)
        if body is None:
            n_errors += 1
            errors.append({
                "idx": idx, "error": "record incompleto: "
                "kind/scope mancanti", "record": rec})
            continue
        owner_name = body.pop("_owner_name", "")
        # owner_id can also be provided explicitly
        if rec.get("owner_id") is not None:
            try:
                body["owner_id"] = int(rec["owner_id"])
            except Exception:
                pass
        if "owner_id" not in body and owner_name:
            owner_id = _resolve_owner_by_name(db, body["scope"],
                                                owner_name)
            if owner_id is None:
                n_errors += 1
                errors.append({
                    "idx": idx,
                    "error": (f"owner_name {owner_name!r} non "
                               f"trovato per scope={body['scope']}"),
                    "record": rec,
                })
                continue
            body["owner_id"] = owner_id
        if "owner_id" not in body:
            n_errors += 1
            errors.append({
                "idx": idx,
                "error": ("ne' owner_id ne' owner_name forniti, e "
                           "il vincolo non e' globale"),
                "record": rec,
            })
            continue
        out = _create_constraint_via_dispatcher(db, body)
        if out.get("ok"):
            n_loaded += 1
        else:
            n_errors += 1
            errors.append({"idx": idx, "error": out.get("error"),
                            "record": rec})
    db.commit()
    try:
        from ..utils.ttl_cache import bump_mutation
        bump_mutation()
    except Exception:
        pass
    return {
        "ok": True,
        "filename": file.filename,
        "n_total": len(raw_records),
        "n_loaded": n_loaded,
        "n_errors": n_errors,
        "errors": errors[:50],   # cap to avoid huge response
    }


# ----- Export ---------------------------------------------------------

def _collect_constraints(db: Session) -> list[dict]:
    """Snapshot every constraint kind that the unified API understands,
    in the canonical import-shape."""
    out: list[dict] = []

    def _owner_name_for(scope: str, oid: int) -> str | None:
        if oid is None:
            return None
        if scope == "teacher":
            t = db.get(_models.Teacher, oid)
            if t is None:
                return None
            for f in ("display_name", "name", "nickname"):
                v = getattr(t, f, None)
                if v:
                    return str(v)
            ln = getattr(t, "last_name", "") or ""
            fn = getattr(t, "first_name", "") or ""
            return f"{ln} {fn}".strip() or None
        if scope == "class":
            c = db.get(_models.SchoolClass, oid)
            return getattr(c, "name", None) if c else None
        if scope == "classroom":
            r = db.get(_models.Classroom, oid)
            return getattr(r, "name", None) if r else None
        if scope == "curriculum":
            c = db.get(_models.Curriculum, oid)
            return getattr(c, "name", None) if c else None
        return None

    # LogicalUnavailability (teacher | class | classroom)
    for r in db.query(_models.LogicalUnavailability).all():
        scope = (r.entity_type or "").lower()
        out.append({
            "kind": "logical",
            "scope": scope,
            "owner_id": r.entity_id,
            "owner_name": _owner_name_for(scope, r.entity_id),
            "level": r.kind or ("hard" if r.is_hard else "soft"),
            "expression": r.expression,
            "soft_penalty": r.soft_penalty,
            "description": None,
        })

    # CurriculumLogicalConstraint
    for r in db.query(_models.CurriculumLogicalConstraint).all():
        out.append({
            "kind": "logical",
            "scope": "curriculum",
            "owner_id": r.curriculum_id,
            "owner_name": _owner_name_for("curriculum", r.curriculum_id),
            "level": r.kind or ("hard" if r.is_hard else "soft"),
            "expression": r.expression,
            "soft_penalty": r.soft_penalty,
            "description": r.label,
            "year_filter": r.year_filter,
        })

    # Matrix slot — TeacherUnavailability / ClassUnavailability /
    # ClassroomUnavailability
    for Model, scope in (
        (_models.TeacherUnavailability, "teacher"),
        (_models.ClassUnavailability, "class"),
        (_models.ClassroomUnavailability, "classroom"),
    ):
        for r in db.query(Model).all():
            owner_id = (getattr(r, "teacher_id", None)
                        or getattr(r, "class_id", None)
                        or getattr(r, "classroom_id", None))
            out.append({
                "kind": "matrix_slot",
                "scope": scope,
                "owner_id": owner_id,
                "owner_name": _owner_name_for(scope, owner_id),
                "level": getattr(r, "state", "hard") or "hard",
                "day": getattr(r, "day", None),
                "hour": getattr(r, "hour", None),
                "soft_penalty": getattr(r, "soft_penalty", None),
                "description": getattr(r, "reason", None),
            })

    # CoTeachingRule
    for r in db.query(_models.CoTeachingRule).all():
        out.append({
            "kind": "coteach",
            "scope": "class",
            "owner_id": r.class_id,
            "owner_name": _owner_name_for("class", r.class_id),
            "level": "hard" if r.required else "soft",
            "subject": r.subject,
            "n_teachers": getattr(r, "n_teachers", None),
            "soft_penalty": int(getattr(r, "weight", 0) or 0),
            "description": None,
        })

    # ClassroomSubjectPreference / TeacherClassroomPreference
    try:
        for r in db.query(_models.ClassroomSubjectPreference).all():
            out.append({
                "kind": "room_pref",
                "scope": "subject_room",
                "owner_id": getattr(r, "classroom_id", None),
                "owner_name": _owner_name_for(
                    "classroom", getattr(r, "classroom_id", None)),
                "level": getattr(r, "state", "preferred"),
                "subject": getattr(r, "subject", None),
                "soft_penalty": getattr(r, "soft_penalty", None),
                "description": None,
            })
    except Exception:
        pass
    try:
        for r in db.query(_models.TeacherClassroomPreference).all():
            out.append({
                "kind": "room_pref",
                "scope": "teacher_room",
                "owner_id": getattr(r, "teacher_id", None),
                "owner_name": _owner_name_for(
                    "teacher", getattr(r, "teacher_id", None)),
                "level": getattr(r, "state", "preferred"),
                "soft_penalty": getattr(r, "soft_penalty", None),
                "description": None,
            })
    except Exception:
        pass

    return out


@router.get("/constraints/export")
def constraints_export(
    format: str = Query("json", pattern="^(json|xlsx)$"),
    db: Session = Depends(get_db),
):
    """Download all constraints in the DB as JSON or xlsx.
    The file shape is the same accepted by /import-file, so it's a
    round-trip."""
    rows = _collect_constraints(db)
    ts = dt.datetime.now().strftime("%Y%m%d_%H%M%S")
    if format == "json":
        body = json.dumps(rows, indent=2, ensure_ascii=False).encode("utf-8")
        fn = f"pitantum_constraints_{ts}.json"
        return StreamingResponse(
            io.BytesIO(body), media_type="application/json",
            headers={"Content-Disposition":
                     f'attachment; filename="{fn}"'},
        )
    # xlsx
    try:
        from openpyxl import Workbook
    except Exception:
        raise HTTPException(
            500, "openpyxl non installato; impossibile produrre xlsx.")
    wb = Workbook()
    ws = wb.active
    ws.title = "constraints"
    headers = ["kind", "scope", "owner_id", "owner_name", "level",
               "expression", "subject", "day", "hour", "year_filter",
               "n_teachers", "soft_penalty", "description"]
    ws.append(headers)
    for r in rows:
        ws.append([r.get(h) for h in headers])
    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    fn = f"pitantum_constraints_{ts}.xlsx"
    return StreamingResponse(
        out, media_type=("application/vnd.openxmlformats-officedocument"
                          ".spreadsheetml.sheet"),
        headers={"Content-Disposition":
                 f'attachment; filename="{fn}"'},
    )


# ----- Delete-all -----------------------------------------------------

@router.delete("/constraints/all")
def constraints_delete_all(
    confirm: bool = Query(
        False,
        description="Must be true to actually wipe the constraint tables."),
    db: Session = Depends(get_db),
):
    """Wipe every constraint table. Requires `?confirm=true`. Returns
    the per-table delete counts."""
    if not confirm:
        raise HTTPException(
            400,
            "richiede ?confirm=true (doppia conferma da UI).")
    counts: dict[str, int] = {}
    for Model, key in (
        (_models.LogicalUnavailability, "logical_unavailabilities"),
        (_models.CurriculumLogicalConstraint,
            "curriculum_logical_constraints"),
        (_models.TeacherUnavailability, "teacher_unavailabilities"),
        (_models.ClassUnavailability, "class_unavailabilities"),
        (_models.ClassroomUnavailability, "classroom_unavailabilities"),
        (_models.CoTeachingRule, "coteaching_rules"),
    ):
        try:
            n = db.query(Model).delete(synchronize_session=False)
            counts[key] = int(n or 0)
        except Exception as e:
            counts[key] = -1
            counts[f"{key}__error"] = str(e)
    for Model, key in (
        (_models.ClassroomSubjectPreference,
            "classroom_subject_preferences"),
        (_models.TeacherClassroomPreference,
            "teacher_classroom_preferences"),
    ):
        try:
            n = db.query(Model).delete(synchronize_session=False)
            counts[key] = int(n or 0)
        except Exception:
            counts[key] = 0
    db.commit()
    try:
        from ..utils.ttl_cache import bump_mutation
        bump_mutation()
    except Exception:
        pass
    total = sum(v for v in counts.values()
                 if isinstance(v, int) and v >= 0)
    return {"ok": True, "n_deleted": total, "by_table": counts}
