"""Unified Excel / CSV bulk-import endpoint.

POST /api/import/{entity}
  multipart/form-data:
    file:        the .xlsx / .xls / .csv / .tsv to import
    mode:        'upsert' (default) | 'replace' | 'append'
    sheet:       optional sheet name (xlsx only); first sheet by default
  -> ImportReport JSON

GET  /api/import/{entity}/template
  -> downloads a sample .xlsx with the expected headers and one example row.

The {entity} field is one of:
  teachers | subjects | classes | classrooms | curricula | students |
  groups | assignments | subject_group_weights

Each entity has its own column schema documented in
webui/docs/import_format.md. Headers are matched case-insensitively after
trimming whitespace; underscores and spaces are interchangeable.

The endpoint is intentionally permissive: it accepts both Italian and
English header names (e.g. 'Cognome'/'last_name', 'Materia'/'subject').
"""
from __future__ import annotations

import csv
import io
import re
from typing import Any, Iterable

from fastapi import APIRouter, Depends, File, Form, HTTPException, UploadFile
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session

from .. import models, schemas
from ..db import get_db

router = APIRouter(prefix="/api/import", tags=["import"])


# ---------- spreadsheet parsing helpers ----------


def _normalize_header(h: Any) -> str:
    if h is None:
        return ""
    return re.sub(r"[\s_-]+", "_", str(h).strip().lower())


def _read_xlsx(content: bytes, sheet: str | None
               ) -> tuple[list[str], list[list[Any]]]:
    """Read a sheet from a xlsx workbook. When sheet is None and a
    'Dati' sheet exists (multi-sheet templates produced by GET
    /api/import/{entity}/template), prefer that one over the default
    first-sheet behaviour.
    """
    try:
        from openpyxl import load_workbook
    except ImportError as e:  # pragma: no cover
        raise HTTPException(500, f"openpyxl non disponibile: {e}")
    wb = load_workbook(io.BytesIO(content), data_only=True, read_only=True)
    if sheet:
        if sheet not in wb.sheetnames:
            raise HTTPException(
                400, f"foglio '{sheet}' non trovato; disponibili: "
                     f"{wb.sheetnames}"
            )
        ws = wb[sheet]
    else:
        # Prefer 'Dati' if present (templates), else first sheet
        if "Dati" in wb.sheetnames:
            ws = wb["Dati"]
        else:
            ws = wb.worksheets[0]
    rows = ws.iter_rows(values_only=True)
    try:
        header = list(next(rows))
    except StopIteration:
        return [], []
    header = [_normalize_header(h) for h in header]
    body = []
    for r in rows:
        if r is None:
            continue
        body.append(list(r))
    return header, body


def _read_csv(content: bytes) -> tuple[list[str], list[list[Any]]]:
    text = content.decode("utf-8-sig", errors="replace")
    sample = text[:8192]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
    except csv.Error:
        class _D:
            delimiter = "," if "," in sample else ";"
            quotechar = '"'
        dialect = _D()
    reader = csv.reader(io.StringIO(text), dialect=dialect)
    rows = list(reader)
    if not rows:
        return [], []
    header = [_normalize_header(h) for h in rows[0]]
    return header, rows[1:]


def _detect_format(filename: str, content: bytes
                   ) -> tuple[list[str], list[list[Any]]]:
    name = (filename or "").lower()
    if name.endswith(".xlsx") or name.endswith(".xls") or name.endswith(".xlsm"):
        return _read_xlsx(content, None)
    if name.endswith(".csv") or name.endswith(".tsv") or name.endswith(".txt"):
        return _read_csv(content)
    # Sniff: xlsx files start with PK
    if content[:2] == b"PK":
        return _read_xlsx(content, None)
    return _read_csv(content)


def _row_to_dict(header: list[str], row: list[Any]) -> dict[str, Any]:
    out: dict[str, Any] = {}
    for i, h in enumerate(header):
        if not h:
            continue
        v = row[i] if i < len(row) else None
        if isinstance(v, str):
            v = v.strip()
            if v == "":
                v = None
        out[h] = v
    return out


def _is_blank_row(d: dict[str, Any]) -> bool:
    return all(v is None or v == "" for v in d.values())


def _norm_bool(v: Any, default: bool = False) -> bool:
    if v is None or v == "":
        return default
    if isinstance(v, bool):
        return v
    if isinstance(v, (int, float)):
        return bool(v)
    s = str(v).strip().lower()
    return s in ("1", "true", "vero", "si", "y", "yes", "x", "on")


def _norm_int(v: Any, default: int = 0) -> int:
    if v is None or v == "":
        return default
    try:
        return int(float(v))
    except Exception:
        return default


def _norm_float(v: Any, default: float = 0.0) -> float:
    if v is None or v == "":
        return default
    try:
        return float(v)
    except Exception:
        return default


def _pick(d: dict[str, Any], *keys: str) -> Any:
    """Return first non-None value among the given alias keys."""
    for k in keys:
        if k in d and d[k] is not None and d[k] != "":
            return d[k]
    return None


def _split_csv_field(s: Any) -> list[str]:
    if s is None or s == "":
        return []
    return [p.strip() for p in str(s).split(",") if p.strip()]


# ---------- per-entity importers ----------


def _import_teachers(db: Session, rows: Iterable[dict],
                     mode: str) -> schemas.ImportReport:
    rep = schemas.ImportReport(ok=True, entity="teachers")
    if mode == "replace":
        db.query(models.TeacherSubject).delete()
        db.query(models.Teacher).delete()
        db.commit()
    by_name = {t.name: t for t in db.query(models.Teacher).all()}
    for r in rows:
        rep.n_total_rows += 1
        if _is_blank_row(r):
            rep.n_skipped += 1
            continue
        name = _pick(r, "name", "nome", "cognome_nome", "fullname")
        if not name:
            rep.n_skipped += 1
            rep.errors.append(f"riga {rep.n_total_rows}: name mancante")
            continue
        name = str(name)
        existing = by_name.get(name)
        is_update = existing is not None
        t = existing or models.Teacher(name=name)
        t.matricola = _pick(r, "matricola", "code")
        t.group = _pick(r, "group", "gruppo", "classe_di_concorso", "cdc")
        t.max_hours = _norm_int(_pick(r, "max_hours", "max_ore", "ore_max"),
                                t.max_hours or 18)
        t.completion_hours = _norm_int(_pick(r, "completion_hours",
                                             "ore_completamento"), 0)
        t.exemption_hours = _norm_int(_pick(r, "exemption_hours",
                                            "ore_esonero"), 0)
        t.free_day = _pick(r, "free_day", "giorno_libero")
        t.max_consecutive = _norm_int(_pick(r, "max_consecutive",
                                            "max_consecutive_ore"), 5)
        # New fields (added since the original importer)
        gs = _pick(r, "graduatoria_score", "graduatoria",
                   "punteggio_graduatoria")
        if gs not in (None, ""):
            t.graduatoria_score = _norm_float(gs, t.graduatoria_score or 0.0)
        rfdc = _pick(r, "required_free_days_count",
                     "giorni_liberi_richiesti")
        if rfdc not in (None, ""):
            t.required_free_days_count = _norm_int(
                rfdc, getattr(t, "required_free_days_count", 1) or 1
            )
        pfd = _pick(r, "preferred_free_days_json",
                     "preferred_free_days")
        if pfd not in (None, ""):
            t.preferred_free_days_json = str(pfd)
        t.notes = _pick(r, "notes", "note")
        if not is_update:
            db.add(t)
            db.flush()
        # Subjects column: comma-separated list
        subjs = _split_csv_field(_pick(r, "subjects", "materie"))
        if subjs:
            db.query(models.TeacherSubject).filter(
                models.TeacherSubject.teacher_id == t.id
            ).delete()
            for s in subjs:
                db.add(models.TeacherSubject(teacher_id=t.id, subject=s))
        if is_update:
            rep.n_updated += 1
        else:
            rep.n_inserted += 1
        by_name[name] = t
    db.commit()
    return rep


def _import_subjects(db: Session, rows: Iterable[dict],
                     mode: str) -> schemas.ImportReport:
    rep = schemas.ImportReport(ok=True, entity="subjects")
    if mode == "replace":
        db.query(models.Subject).delete()
        db.commit()
    by_name = {s.name: s for s in db.query(models.Subject).all()}
    for r in rows:
        rep.n_total_rows += 1
        if _is_blank_row(r):
            rep.n_skipped += 1
            continue
        name = _pick(r, "name", "nome", "materia", "subject")
        if not name:
            rep.n_skipped += 1
            rep.errors.append(f"riga {rep.n_total_rows}: name mancante")
            continue
        existing = by_name.get(str(name))
        is_update = existing is not None
        s = existing or models.Subject(name=str(name))
        s.pretty_name = _pick(r, "pretty_name", "nome_esteso") or name
        s.notes = _pick(r, "notes", "note")
        s.distribute_days_weight = _norm_float(_pick(r, "distribute_days_weight",
                                                     "peso_distribuzione"), 0.0)
        s.dual_hours_weight = _norm_float(_pick(r, "dual_hours_weight",
                                                "peso_doppia"), 0.0)
        s.no_sixth_hour_weight = _norm_float(_pick(r, "no_sixth_hour_weight",
                                                   "peso_no_6ora"), 0.0)
        if not is_update:
            db.add(s)
        if is_update:
            rep.n_updated += 1
        else:
            rep.n_inserted += 1
        by_name[s.name] = s
    db.commit()
    return rep


def _import_classes(db: Session, rows: Iterable[dict],
                    mode: str) -> schemas.ImportReport:
    rep = schemas.ImportReport(ok=True, entity="classes")
    if mode == "replace":
        db.query(models.ClassSubject).delete()
        db.query(models.SchoolClass).delete()
        db.commit()
    by_name = {c.name: c for c in db.query(models.SchoolClass).all()}
    curr_by_code = {c.code: c for c in db.query(models.Curriculum).all()}
    # Group rows by class name so multiple subjects under same class merge
    grouped: dict[str, list[dict]] = {}
    base_rows: dict[str, dict] = {}
    for r in rows:
        rep.n_total_rows += 1
        if _is_blank_row(r):
            rep.n_skipped += 1
            continue
        name = _pick(r, "name", "nome", "classe", "class")
        if not name:
            rep.n_skipped += 1
            rep.errors.append(f"riga {rep.n_total_rows}: name mancante")
            continue
        grouped.setdefault(str(name), []).append(r)
        base_rows.setdefault(str(name), r)
    for name, rs in grouped.items():
        r0 = base_rows[name]
        existing = by_name.get(name)
        is_update = existing is not None
        c = existing or models.SchoolClass(name=name)
        c.year = _norm_int(_pick(r0, "year", "anno"), c.year or 1)
        c.section = _pick(r0, "section", "sezione")
        curr = _pick(r0, "curriculum", "indirizzo", "curriculum_code")
        c.curriculum = str(curr) if curr else None
        if curr and str(curr) in curr_by_code:
            c.curriculum_id = curr_by_code[str(curr)].id
        c.n_students = _norm_int(_pick(r0, "n_students", "studenti"),
                                 c.n_students or 22)
        # New: max_hours_per_day, preferred_free_days_json,
        # required_free_days_count
        mhpd = _pick(r0, "max_hours_per_day", "max_ore_giorno")
        if mhpd not in (None, ""):
            c.max_hours_per_day = _norm_int(
                mhpd, getattr(c, "max_hours_per_day", 5) or 5
            )
        rfdc = _pick(r0, "required_free_days_count",
                     "giorni_liberi_richiesti")
        if rfdc not in (None, ""):
            c.required_free_days_count = _norm_int(
                rfdc, getattr(c, "required_free_days_count", 0) or 0
            )
        pfd = _pick(r0, "preferred_free_days_json",
                     "preferred_free_days")
        if pfd not in (None, ""):
            c.preferred_free_days_json = str(pfd)
        c.notes = _pick(r0, "notes", "note")
        if not is_update:
            db.add(c)
            db.flush()
        # subjects: collect rows with subject + hours
        subj_pairs: list[tuple[str, int]] = []
        for r in rs:
            subj = _pick(r, "subject", "materia")
            if subj:
                hh = _norm_int(_pick(r, "hours_per_week", "ore",
                                     "ore_settimanali"), 0)
                subj_pairs.append((str(subj), hh))
        if subj_pairs:
            db.query(models.ClassSubject).filter(
                models.ClassSubject.class_id == c.id
            ).delete()
            for s, h in subj_pairs:
                db.add(models.ClassSubject(class_id=c.id,
                                           subject=s, hours_per_week=h))
        if is_update:
            rep.n_updated += 1
        else:
            rep.n_inserted += 1
        by_name[name] = c
    db.commit()
    return rep


def _import_classrooms(db: Session, rows: Iterable[dict],
                       mode: str) -> schemas.ImportReport:
    rep = schemas.ImportReport(ok=True, entity="classrooms")
    if mode == "replace":
        db.query(models.Classroom).delete()
        db.commit()
    by_name = {r.name: r for r in db.query(models.Classroom).all()}
    for r in rows:
        rep.n_total_rows += 1
        if _is_blank_row(r):
            rep.n_skipped += 1
            continue
        name = _pick(r, "name", "nome", "aula", "classroom")
        if not name:
            rep.n_skipped += 1
            rep.errors.append(f"riga {rep.n_total_rows}: name mancante")
            continue
        existing = by_name.get(str(name))
        is_update = existing is not None
        room = existing or models.Classroom(name=str(name))
        room.kind = _pick(r, "kind", "tipo") or "standard"
        room.capacity = _norm_int(_pick(r, "capacity", "capienza"),
                                  room.capacity or 30)
        room.multi_class = _norm_bool(_pick(r, "multi_class", "multi_classe"),
                                      room.multi_class)
        room.multi_class_max = _norm_int(_pick(r, "multi_class_max",
                                               "multi_classe_max"), 1)
        room.notes = _pick(r, "notes", "note")
        if not is_update:
            db.add(room)
            db.flush()
        # ----- Tags column (comma-separated) -----
        tag_field = _pick(r, "tags", "tag", "etichette")
        if tag_field is not None:
            wanted = {
                t.strip().lower()
                for t in _split_csv_field(tag_field)
                if t and t.strip()
            }
            existing_tags = {
                t.name: t for t in db.query(models.ClassroomTag).filter(
                    models.ClassroomTag.name.in_(wanted)
                ).all()
            } if wanted else {}
            for n in wanted - set(existing_tags):
                new_tag = models.ClassroomTag(name=n)
                db.add(new_tag)
                db.flush()
                existing_tags[n] = new_tag
            db.query(models.ClassroomTagAssignment).filter(
                models.ClassroomTagAssignment.classroom_id == room.id
            ).delete()
            for n in wanted:
                db.add(models.ClassroomTagAssignment(
                    classroom_id=room.id, tag_id=existing_tags[n].id,
                ))
        if is_update:
            rep.n_updated += 1
        else:
            rep.n_inserted += 1
        by_name[room.name] = room
    db.commit()
    return rep


def _import_curricula(db: Session, rows: Iterable[dict],
                      mode: str) -> schemas.ImportReport:
    rep = schemas.ImportReport(ok=True, entity="curricula")
    if mode == "replace":
        db.query(models.CurriculumSubjectHours).delete()
        db.query(models.Curriculum).delete()
        db.commit()
    by_code = {c.code: c for c in db.query(models.Curriculum).all()}
    grouped: dict[str, list[dict]] = {}
    base_rows: dict[str, dict] = {}
    for r in rows:
        rep.n_total_rows += 1
        if _is_blank_row(r):
            rep.n_skipped += 1
            continue
        code = _pick(r, "code", "codice", "curriculum",
                     "indirizzo", "curriculum_code")
        if not code:
            rep.n_skipped += 1
            rep.errors.append(f"riga {rep.n_total_rows}: code mancante")
            continue
        grouped.setdefault(str(code), []).append(r)
        base_rows.setdefault(str(code), r)
    for code, rs in grouped.items():
        r0 = base_rows[code]
        existing = by_code.get(code)
        is_update = existing is not None
        c = existing or models.Curriculum(code=code, name=code)
        nm = _pick(r0, "name", "nome")
        if nm:
            c.name = str(nm)
        c.description = _pick(r0, "description", "descrizione")
        c.notes = _pick(r0, "notes", "note")
        c.score = _norm_int(_pick(r0, "score", "punteggio"), c.score or 1)
        if not is_update:
            db.add(c)
            db.flush()
        # hours: each row carries year + subject + hours
        hours_triples: list[tuple[int, str, int]] = []
        for r in rs:
            year = _pick(r, "year", "anno")
            subj = _pick(r, "subject", "materia")
            if year is None or not subj:
                continue
            yi = _norm_int(year, 0)
            hh = _norm_int(_pick(r, "hours_per_week", "ore",
                                 "ore_settimanali"), 0)
            if yi >= 1:
                hours_triples.append((yi, str(subj), hh))
        if hours_triples:
            db.query(models.CurriculumSubjectHours).filter(
                models.CurriculumSubjectHours.curriculum_id == c.id
            ).delete()
            seen = set()
            for yi, s, hh in hours_triples:
                key = (yi, s)
                if key in seen:
                    continue
                seen.add(key)
                db.add(models.CurriculumSubjectHours(
                    curriculum_id=c.id, year=yi,
                    subject=s, hours_per_week=hh
                ))
        if is_update:
            rep.n_updated += 1
        else:
            rep.n_inserted += 1
        by_code[code] = c
    db.commit()
    return rep


def _import_students(db: Session, rows: Iterable[dict],
                     mode: str) -> schemas.ImportReport:
    rep = schemas.ImportReport(ok=True, entity="students")
    if mode == "replace":
        db.query(models.GroupMembership).delete()
        db.query(models.Student).delete()
        db.commit()
    classes_by_name = {c.name: c.id for c in db.query(models.SchoolClass).all()}
    # composite key (last_name, first_name, birth_date) for matching
    existing = {(s.last_name, s.first_name, s.birth_date): s
                for s in db.query(models.Student).all()}
    for r in rows:
        rep.n_total_rows += 1
        if _is_blank_row(r):
            rep.n_skipped += 1
            continue
        last = _pick(r, "last_name", "cognome")
        first = _pick(r, "first_name", "nome")
        if not last or not first:
            rep.n_skipped += 1
            rep.errors.append(f"riga {rep.n_total_rows}: cognome/nome mancante")
            continue
        bd = _pick(r, "birth_date", "data_nascita")
        # normalize date if datetime/string
        import datetime as _dt
        if isinstance(bd, _dt.datetime):
            bd = bd.date()
        elif isinstance(bd, str):
            try:
                bd = _dt.date.fromisoformat(bd[:10])
            except Exception:
                bd = None
        key = (str(last), str(first), bd)
        s = existing.get(key)
        is_update = s is not None
        if s is None:
            s = models.Student(last_name=str(last), first_name=str(first),
                               birth_date=bd)
            db.add(s)
        s.last_name = str(last)
        s.first_name = str(first)
        s.birth_date = bd
        s.gender = _pick(r, "gender", "sesso")
        s.email = _pick(r, "email")
        s.student_code = _pick(r, "student_code", "matricola", "codice")
        cname = _pick(r, "class_name", "classe", "class")
        s.class_id = classes_by_name.get(str(cname)) if cname else None
        s.notes = _pick(r, "notes", "note")
        # ----- Tags (created on the fly, lower-cased) -----
        tag_field = _pick(r, "tags", "tag", "etichette")
        if tag_field is not None:
            db.flush()
            wanted = {
                t.strip().lower()
                for t in _split_csv_field(tag_field)
                if t and t.strip()
            }
            # ensure all tag rows exist
            existing_tags = {
                t.name: t for t in db.query(models.StudentTag).filter(
                    models.StudentTag.name.in_(wanted)
                ).all()
            } if wanted else {}
            for n in wanted - set(existing_tags):
                new_tag = models.StudentTag(name=n)
                db.add(new_tag)
                db.flush()
                existing_tags[n] = new_tag
            # rewrite assignments
            db.query(models.StudentTagAssignment).filter(
                models.StudentTagAssignment.student_id == s.id
            ).delete()
            for n in wanted:
                db.add(models.StudentTagAssignment(
                    student_id=s.id, tag_id=existing_tags[n].id,
                ))
        if is_update:
            rep.n_updated += 1
        else:
            rep.n_inserted += 1
        existing[key] = s
    db.commit()
    return rep


def _import_groups(db: Session, rows: Iterable[dict],
                   mode: str) -> schemas.ImportReport:
    """Groups import format: each row is a (group_name, kind, subject,
    hours_per_week, student_codes_csv) tuple; rows with the same group_name
    are merged."""
    rep = schemas.ImportReport(ok=True, entity="groups")
    if mode == "replace":
        db.query(models.GroupMembership).delete()
        db.query(models.GroupSubjectHours).delete()
        db.query(models.StudyGroup).delete()
        db.commit()
    by_name = {g.name: g for g in db.query(models.StudyGroup).all()}
    students_by_code = {s.student_code: s for s in db.query(models.Student).all()
                        if s.student_code}
    grouped: dict[str, list[dict]] = {}
    for r in rows:
        rep.n_total_rows += 1
        if _is_blank_row(r):
            rep.n_skipped += 1
            continue
        name = _pick(r, "name", "nome", "group", "gruppo")
        if not name:
            rep.n_skipped += 1
            rep.errors.append(f"riga {rep.n_total_rows}: name mancante")
            continue
        grouped.setdefault(str(name), []).append(r)
    for name, rs in grouped.items():
        r0 = rs[0]
        existing = by_name.get(name)
        is_update = existing is not None
        g = existing or models.StudyGroup(name=name)
        kind = _pick(r0, "kind", "tipo")
        g.kind = str(kind) if kind else (g.kind or "splitting")
        g.description = _pick(r0, "description", "descrizione")
        g.notes = _pick(r0, "notes", "note")
        if not is_update:
            db.add(g)
            db.flush()
        # subject hours
        seen_subj = set()
        for r in rs:
            sj = _pick(r, "subject", "materia")
            if sj and sj not in seen_subj:
                seen_subj.add(sj)
                hh = _norm_int(_pick(r, "hours_per_week", "ore"), 1)
                row = db.query(models.GroupSubjectHours).filter(
                    models.GroupSubjectHours.group_id == g.id,
                    models.GroupSubjectHours.subject == sj,
                ).first()
                if row is None:
                    db.add(models.GroupSubjectHours(
                        group_id=g.id, subject=str(sj),
                        hours_per_week=int(hh)
                    ))
                else:
                    row.hours_per_week = int(hh)
        # members from comma-separated student codes
        codes: set[str] = set()
        for r in rs:
            for c in _split_csv_field(_pick(r, "student_codes", "matricole",
                                            "students")):
                codes.add(c)
        if codes:
            db.query(models.GroupMembership).filter(
                models.GroupMembership.group_id == g.id
            ).delete()
            for code in codes:
                st = students_by_code.get(code)
                if st is None:
                    rep.errors.append(f"gruppo {name}: studente "
                                      f"con codice '{code}' non trovato")
                    continue
                db.add(models.GroupMembership(
                    group_id=g.id, student_id=st.id
                ))
        if is_update:
            rep.n_updated += 1
        else:
            rep.n_inserted += 1
        by_name[name] = g
    db.commit()
    return rep


_IMPORTERS = {
    "teachers": _import_teachers,
    "subjects": _import_subjects,
    "classes": _import_classes,
    "classrooms": _import_classrooms,
    "curricula": _import_curricula,
    "students": _import_students,
    "groups": _import_groups,
}


# ---------- HTTP endpoints ----------


@router.post("/{entity}", response_model=schemas.ImportReport)
async def do_import(entity: str,
                    file: UploadFile = File(...),
                    mode: str = Form("upsert"),
                    db: Session = Depends(get_db)):
    if entity not in _IMPORTERS:
        raise HTTPException(404, f"entity '{entity}' non supportata")
    if mode not in ("upsert", "replace", "append"):
        raise HTTPException(400, "mode deve essere upsert/replace/append")
    content = await file.read()
    if not content:
        raise HTTPException(400, "file vuoto")
    try:
        header, body = _detect_format(file.filename or "", content)
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(400, f"errore lettura file: {e}")
    if not header:
        raise HTTPException(400, "header non trovato (prima riga vuota?)")
    rows = [_row_to_dict(header, r) for r in body]
    importer = _IMPORTERS[entity]
    return importer(db, rows, mode)


# ---------- template generator ----------


_TEMPLATES: dict[str, list[tuple[str, Any]]] = {
    "teachers": [
        ("name", "Mario Rossi"),
        ("matricola", "MR123"),
        ("group", "A050"),
        ("max_hours", 18),
        ("free_day", "Saturday"),
        ("subjects", "Matematica,Fisica"),
        ("graduatoria_score", 0.0),
        ("required_free_days_count", 1),
        ("preferred_free_days_json", ""),
        ("notes", ""),
    ],
    "subjects": [
        ("name", "Matematica"),
        ("pretty_name", "Matematica"),
        ("distribute_days_weight", 1.0),
        ("dual_hours_weight", 0.5),
        ("no_sixth_hour_weight", 0.0),
        ("notes", ""),
    ],
    "classes": [
        ("name", "1A"),
        ("year", 1),
        ("section", "A"),
        ("curriculum", "Scientifico"),
        ("n_students", 22),
        ("subject", "Matematica"),
        ("hours_per_week", 5),
        ("max_hours_per_day", 5),
        ("required_free_days_count", 0),
        ("preferred_free_days_json", ""),
        ("notes", ""),
    ],
    "classrooms": [
        ("name", "Aula 12"),
        ("kind", "standard"),
        ("capacity", 30),
        ("multi_class", "no"),
        ("multi_class_max", 1),
        ("tags", "matematica,scientifico"),
        ("notes", ""),
    ],
    "curricula": [
        ("code", "Scientifico"),
        ("name", "Liceo Scientifico"),
        ("description", "Indirizzo scientifico tradizionale"),
        ("score", 1),
        ("year", 1),
        ("subject", "Matematica"),
        ("hours_per_week", 5),
    ],
    "students": [
        ("last_name", "Rossi"),
        ("first_name", "Anna"),
        ("birth_date", "2010-05-12"),
        ("gender", "F"),
        ("email", ""),
        ("student_code", "S001"),
        ("class_name", "1A"),
        ("tags", "BES,debito_matematica_4"),
        ("notes", ""),
    ],
    "groups": [
        ("name", "Spagnolo 1A+1B"),
        ("kind", "language"),
        ("description", "Gruppo seconda lingua"),
        ("subject", "Spagnolo"),
        ("hours_per_week", 3),
        ("student_codes", "S001,S002,S003"),
        ("notes", ""),
    ],
}


# Per-column hints displayed in the "Istruzioni" sheet of the
# downloadable template. Keys NOT listed here use a generic
# "campo libero" message.
_TEMPLATE_HINTS: dict[str, dict[str, str]] = {
    "teachers": {
        "name": "OBBLIGATORIO -- nome univoco del docente (chiave).",
        "matricola": "Codice esterno (matricola). Opzionale.",
        "group": "Classe di concorso (es. A050, A026).",
        "max_hours": "Ore settimanali standard (es. 18).",
        "free_day": "Giorno libero legacy (Monday..Saturday). "
                    "Convertito in HARD-cells.",
        "subjects": "Lista materie separate da virgola. Esempio: "
                    "'Matematica,Fisica'.",
        "graduatoria_score": "Float; usato dal modulo supplenze per "
                             "ordinare le candidature.",
        "required_free_days_count": "HARD: numero esatto di giorni "
                                     "liberi a settimana (CCNL=1).",
        "preferred_free_days_json": "Opzionale -- JSON [{day:1..6,"
                                     "is_hard:bool,soft_penalty:int}]. "
                                     "Lascia vuoto per default.",
        "notes": "Annotazioni libere.",
    },
    "classes": {
        "name": "OBBLIGATORIO -- nome classe (es. 1A_Scientifico).",
        "year": "Anno (1..5).",
        "section": "Sezione (A,B,C,...). Opzionale.",
        "curriculum": "Codice indirizzo (vedi /curricula).",
        "n_students": "Numero studenti.",
        "subject": "Una riga per (classe,materia): elenca qui la "
                   "materia di questa riga.",
        "hours_per_week": "Ore settimanali della materia indicata.",
        "max_hours_per_day": "HARD: massimo ore al giorno (default 5).",
        "required_free_days_count": "HARD: numero esatto di giorni "
                                     "liberi (default 0 per classi).",
        "preferred_free_days_json": "JSON come per i docenti.",
    },
    "classrooms": {
        "name": "OBBLIGATORIO -- nome aula (es. 'Aula 12', "
                "'Lab Fisica 1').",
        "kind": "Uno di: standard / lab_chimica / lab_fisica / "
                 "lab_informatica / lab_linguistico / palestra / "
                 "biblioteca / aula_speciale.",
        "capacity": "Capienza (intero).",
        "multi_class": "yes/no (palestre+biblioteche multi-classe).",
        "multi_class_max": "Max classi simultanee (HARD).",
        "tags": "Lista tag separati da virgola (lower-case, creati "
                "al volo). Esempio: 'matematica,scientifico'.",
    },
    "curricula": {
        "code": "OBBLIGATORIO -- codice macchina (Scientifico, "
                 "Linguistico, Itis...).",
        "name": "Nome visibile.",
        "score": "Peso normalizzato dal motore mock.",
        "year": "Anno (1..5).",
        "subject": "Materia (una riga per (curriculum,year,subject)).",
        "hours_per_week": "Ore settimanali del monte ore.",
    },
    "students": {
        "last_name": "OBBLIGATORIO -- cognome.",
        "first_name": "OBBLIGATORIO -- nome.",
        "birth_date": "Formato ISO (YYYY-MM-DD). Parte della chiave "
                       "univoca insieme a cognome+nome.",
        "gender": "M/F/other (opzionale).",
        "student_code": "Matricola (opzionale).",
        "class_name": "Nome classe di appartenenza (vedi /classes).",
        "tags": "Lista tag separati da virgola (lower-case, creati "
                "al volo). Esempio: 'BES,debito_matematica_4'.",
    },
    "groups": {
        "name": "OBBLIGATORIO -- nome gruppo univoco.",
        "kind": "splitting / language / religion / support / other.",
        "subject": "Una riga per (gruppo,materia).",
        "hours_per_week": "Ore settimanali della materia.",
        "student_codes": "Lista student_code separati da virgola.",
    },
    "subjects": {
        "name": "OBBLIGATORIO -- nome materia (chiave).",
        "pretty_name": "Etichetta visualizzata nell'orario.",
        "distribute_days_weight": "SOFT -- peso per distribuire le ore "
                                   "su giorni diversi.",
        "dual_hours_weight": "SOFT -- peso per favorire ore doppie.",
        "no_sixth_hour_weight": "SOFT -- penalizza la 6a ora.",
    },
}


@router.get("/{entity}/template")
def download_template(entity: str):
    """Download a 3-sheet xlsx template:
       - Istruzioni: per-column hint table + general usage notes
       - Esempi:    a few pre-populated example rows
       - Dati:      the empty sheet the user fills in (this is the
                    sheet POST /api/import/{entity} reads by default)
    """
    if entity not in _TEMPLATES:
        raise HTTPException(404, f"entity '{entity}' non supportata")
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError as e:  # pragma: no cover
        raise HTTPException(500, f"openpyxl non disponibile: {e}")
    rows = _TEMPLATES[entity]
    headers = [k for k, _ in rows]
    sample = [v for _, v in rows]
    hints = _TEMPLATE_HINTS.get(entity, {})

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="2563EB")

    wb = Workbook()
    # 1) Istruzioni
    ws_help = wb.active
    ws_help.title = "Istruzioni"
    ws_help.append(["Campo", "Hint"])
    for cell in ws_help[1]:
        cell.font = header_font
        cell.fill = header_fill
    ws_help.column_dimensions["A"].width = 30
    ws_help.column_dimensions["B"].width = 90
    for k, _ in rows:
        ws_help.append([k, hints.get(k, "(campo libero)")])
    # General notes block
    ws_help.append([])
    ws_help.append(["NOTE GENERALI"])
    for n in [
        "L'import accetta i sinonimi italiani dei campi (es. 'cognome' "
            "<-> 'last_name', 'classe' <-> 'class_name').",
        "Le righe vuote vengono ignorate. Se manca un campo "
            "obbligatorio la riga viene saltata e segnalata in "
            "ImportReport.errors.",
        "I valori di liste (subjects, tags, student_codes) sono separati "
            "da virgola.",
        "Per gli xlsx la prima riga del foglio 'Dati' deve contenere gli "
            "header. Per CSV stesse regole, separatore auto-detected.",
        "Mode possibili: 'upsert' (default), 'replace' (cancella e "
            "ricarica tutto), 'append' (solo inserimenti).",
    ]:
        ws_help.append(["", n])

    # 2) Esempi
    ws_ex = wb.create_sheet("Esempi")
    ws_ex.append(headers)
    for cell in ws_ex[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    ws_ex.append(sample)

    # 3) Dati (empty user-fillable sheet)
    ws_data = wb.create_sheet("Dati")
    ws_data.append(headers)
    for cell in ws_data[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    # Column widths so the user sees the headers
    for i, h in enumerate(headers, start=1):
        ws_data.column_dimensions[
            ws_data.cell(row=1, column=i).column_letter
        ].width = max(14, min(40, len(h) + 4))
        ws_ex.column_dimensions[
            ws_ex.cell(row=1, column=i).column_letter
        ].width = max(14, min(40, len(h) + 4))

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fn = f"template_{entity}.xlsx"
    return StreamingResponse(
        buf, media_type=("application/vnd.openxmlformats-"
                         "officedocument.spreadsheetml.sheet"),
        headers={"Content-Disposition": f'attachment; filename="{fn}"'}
    )
