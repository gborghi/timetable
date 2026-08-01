"""Conversion between the DB model (rich constraints) and the engine
pickles (slim structures consumed by engine/*.py).

Note: a newer single-file SQLite snapshot
(``engine/scripts/data/<profile>/<profile>.sqlite``, built by
``engine.scripts.build_profile_db``) is now the canonical source of
truth for the demo profiles. ``import_profile_sqlite_into_db`` below
wires it into ``import_engine_profile`` so the legacy pickle path
becomes a fallback.

The engine works on three pickle shapes:

    school = {
        'profile':   str,
        'classes':   [{name, year, section, curriculum, subjects: {subj: ore}}],
        'teachers':  [{name, group, max_hours, free_day, weights: {subj:int}}],
        'cconcorsopersubject': {subj: {group: weight}},
        'curriculum_scores':   {curriculum: int}
    }
    profs = {
        teacher_name: {
            'classi':  {class_name: {subject: {'ore': N}}},
            'glibero': [d1, d2, d3],
            'day_capacity': {day: max_hours}
        }
    }
    solution = {(prof, class, subj, day, hour): 0|1}

This module hides those details from the rest of the backend.
"""
from __future__ import annotations

import datetime as dt
import json
import os
import random
from collections import defaultdict
from typing import Any

from sqlalchemy.orm import Session

from . import models

DAY_MAP = {
    "Monday": 1, "Tuesday": 2, "Wednesday": 3,
    "Thursday": 4, "Friday": 5, "Saturday": 6,
    # Italian aliases (mock generator emits English; users can pick Italian)
    "Lunedi": 1, "Lunedi'": 1, "Martedi": 2, "Martedi'": 2,
    "Mercoledi": 3, "Mercoledi'": 3, "Giovedi": 4, "Giovedi'": 4,
    "Venerdi": 5, "Venerdi'": 5, "Sabato": 6,
}
DAY_NAME_IT = {1: "Lun", 2: "Mar", 3: "Mer", 4: "Gio", 5: "Ven", 6: "Sab"}


# ---------- DB -> pickles ----------


def school_dict_from_db(db: Session) -> dict[str, Any]:
    """Build the dict that big_mock_school.py would serialize as
    school_<profile>.pkl."""
    # Build a code -> Curriculum map so we can fall back to the indirizzo
    # monte-ore when a class has no per-class subject overrides yet.
    curr_by_id = {c.id: c for c in db.query(models.Curriculum).all()}
    curr_by_code = {c.code: c for c in curr_by_id.values()}
    curr_hours: dict[int, dict[int, dict[str, int]]] = {}
    for h in db.query(models.CurriculumSubjectHours).all():
        curr_hours.setdefault(h.curriculum_id, {})\
            .setdefault(h.year, {})[h.subject] = h.hours_per_week

    classes_dump = []
    for cl in db.query(models.SchoolClass).order_by(models.SchoolClass.name).all():
        # subjects: use per-class rows when present; otherwise fall back to
        # the curriculum monte-ore for the relevant year
        subjects = {s.subject: s.hours_per_week for s in cl.subjects}
        if not subjects and cl.curriculum_id and cl.curriculum_id in curr_hours:
            subjects = dict(curr_hours[cl.curriculum_id].get(cl.year, {}))
        curr_code = None
        if cl.curriculum_id and cl.curriculum_id in curr_by_id:
            curr_code = curr_by_id[cl.curriculum_id].code
        elif cl.curriculum:
            curr_code = cl.curriculum
        classes_dump.append({
            "name": cl.name,
            "year": cl.year,
            "section": cl.section,
            "curriculum": curr_code,
            "subjects": subjects,
        })
    teachers_dump = []
    for t in db.query(models.Teacher).order_by(models.Teacher.name).all():
        cls_prefs = [
            {"class_name": p.class_name, "state": p.state,
             "soft_penalty": p.soft_penalty}
            for p in t.class_preferences
        ]
        cur_prefs = [
            {"curriculum_code": p.curriculum_code, "state": p.state,
             "soft_penalty": p.soft_penalty}
            for p in t.curriculum_preferences
        ]
        teachers_dump.append({
            "name": t.name,
            "group": t.group or "",
            "max_hours": t.max_hours,
            "free_day": t.free_day or "Saturday",
            "weights": {ts.subject: 1 for ts in t.subjects},
            "graduatoria_score": t.graduatoria_score,
            "class_preferences": cls_prefs,
            "curriculum_preferences": cur_prefs,
        })
    cconc = defaultdict(dict)
    for row in db.query(models.SubjectGroupWeight).all():
        cconc[row.subject][row.group_name] = row.weight
    # curriculum_scores: prefer the score column from curricula, fall back
    # to first-segment of legacy `curriculum` string for unmapped classes.
    curriculum_scores: dict[str, int] = {}
    for c in curr_by_id.values():
        curriculum_scores[c.code] = int(c.score or 1)
    for cl in db.query(models.SchoolClass).all():
        if cl.curriculum and cl.curriculum.split("_")[0] not in curriculum_scores:
            curriculum_scores[cl.curriculum.split("_")[0]] = 1
    # curricula: full grid + score, included so the engine can know about
    # indirizzi without having to re-derive them from class names.
    curricula_dump = []
    for c in curr_by_id.values():
        curricula_dump.append({
            "code": c.code,
            "name": c.name,
            "score": int(c.score or 1),
            "hours": {y: dict(ss) for y, ss in
                      curr_hours.get(c.id, {}).items()},
        })
    # students + study groups (not strictly needed by the solver yet, but
    # piped through so future engine versions can see them in one shot)
    students_dump = []
    classes_by_id = {cl.id: cl for cl in
                     db.query(models.SchoolClass).all()}
    for st in db.query(models.Student).order_by(
            models.Student.last_name, models.Student.first_name).all():
        students_dump.append({
            "id": st.id,
            "last_name": st.last_name,
            "first_name": st.first_name,
            "class_name": (classes_by_id[st.class_id].name
                           if st.class_id and st.class_id in classes_by_id
                           else None),
            "student_code": st.student_code,
        })
    groups_dump = []
    for g in db.query(models.StudyGroup).order_by(models.StudyGroup.name).all():
        member_ids = [m.student_id for m in g.members]
        groups_dump.append({
            "name": g.name,
            "kind": g.kind,
            "student_ids": member_ids,
            "subjects": {h.subject: h.hours_per_week
                         for h in g.subject_hours},
        })
    return {
        "profile": "webui",
        "classes": classes_dump,
        "teachers": teachers_dump,
        "cconcorsopersubject": dict(cconc),
        "curriculum_scores": curriculum_scores,
        "curricula": curricula_dump,
        "students": students_dump,
        "groups": groups_dump,
    }


def day_capacity_by_teacher(db: Session) -> dict[int, dict[int, int]]:
    """teacher_id -> {legacy day number: max hours workable that day}.

    Only restricted days appear: a day with no HARD constraint is
    absent, meaning "the day's full capacity". The ceiling is

        configured hours of the day - HARD unavailable hours

    from two sources:
      * ``TeacherMandatoryFreeDay`` rows -> capacity 0;
      * ``TeacherUnavailability`` cells with ``state == "hard"``, which
        is how a school actually enters a part-time or a completamento
        su altra scuola -- cell by cell.

    This is the form in which per-cell unavailability can reach Phase A
    at all. Those cells name an hour and Phase A has no hour dimension,
    but the *count* of blocked hours bounds ``prof_day_load``, which
    Phase A does speak. Solvers that don't preload the DSL stream from
    a DB session read this off the profs dict -- the same channel
    ``min_free_days`` uses.
    """
    hours_by_day: dict[int, set[int]] = {}
    for d in db.query(models.WorkingDay).all():
        if not d.is_active or d.legacy_day_number is None:
            continue
        hours = {int(s.legacy_hour_number) for s in (d.slots or [])
                 if s.legacy_hour_number is not None}
        if hours:
            hours_by_day[int(d.legacy_day_number)] = hours
    if not hours_by_day:
        # Pre-Tab-Ore datasets: the legacy 6x6 grid.
        hours_by_day = {d: set(range(8, 14)) for d in range(1, 7)}

    out: dict[int, dict[int, int]] = {}
    hard: dict[int, dict[int, set[int]]] = {}
    for r in db.query(models.TeacherUnavailability).all():
        if r.state == "hard":
            hard.setdefault(r.teacher_id, {})\
                .setdefault(int(r.day), set()).add(int(r.hour))
    for tid, by_day in hard.items():
        for day, blocked_hours in by_day.items():
            configured = hours_by_day.get(day)
            if not configured:
                continue
            blocked = configured & blocked_hours
            if not blocked:
                continue
            out.setdefault(tid, {})[day] = \
                len(configured) - len(blocked)
    # Mandatory free days win over any partial ceiling.
    for r in db.query(models.TeacherMandatoryFreeDay).all():
        out.setdefault(r.teacher_id, {})[int(r.day)] = 0
    return out


def support_class_id(a, students_by_id: dict) -> int | None:
    r"""Which class a sostegno Assignment actually shadows.

    A support teacher is assigned to a *pupil*, so the pupil's current
    class is the authoritative answer and `Assignment.class_id` is only
    a denormalized copy of it. They diverge whenever a pupil is moved
    to another class after the cattedra was created; re-deriving here
    means the support teacher follows the pupil instead of staying
    behind in the old class with nobody to support.

    Returns None when there is nothing to shadow (no pupil and no
    class) -- the caller is expected to skip the row; the preflight
    check is what tells the user about it.
    """
    if getattr(a, "student_id", None) is not None:
        st = students_by_id.get(a.student_id)
        if st is not None and st.class_id is not None:
            return int(st.class_id)
    return a.class_id


def profs_dict_from_db(db: Session) -> dict[str, Any]:
    """Build profs_<profile>.pkl content from the active assignments.

    Task C1 update:
    - Shared coteaching: each member of a coteach group has its own
      Assignment row, so all of them naturally land in `profs` with
      the cattedra's full hours (the principal teacher's hours --
      the n_hours-of-overlap is enforced by the solver via the
      coteach_groups_for_solver helper, NOT by inflating triples here).
    - Potenziamento (class_id is NULL): excluded from `profs` entirely
      because the solver handles them through a separate
      potenziamento_assignments_from_db channel; the standard
      Phase A / Phase B variables are class-bound and don't fit a
      class-less cattedra.
    - Sostegno (is_support=True): included in `profs` as a regular
      triple, under the class of the pupil it follows (see
      support_class_id). The shadow constraint is added separately by
      the solver using support_assignments_from_db.
    """
    rng = random.Random(123)
    days = list(range(1, 7))
    # Use teacher.free_day for the primary free day
    out: dict[str, Any] = {}
    teachers = {t.id: t for t in db.query(models.Teacher).all()}
    classes = {c.id: c for c in db.query(models.SchoolClass).all()}
    studygroups = {sg.id: sg for sg in db.query(models.StudyGroup).all()}
    students = {s.id: s for s in db.query(models.Student).all()}
    for a in db.query(models.Assignment).all():
        if a.is_potenziamento:
            continue
        t = teachers.get(a.teacher_id)
        if t is None:
            continue
        # Task C3: group-targeted Assignments (group_id set) are
        # NOT inserted into profs.classi -- the solver receives them
        # via group_assignments_for_solver as a separate channel.
        # The augmentation logic in cv2.solve_phase_a creates the
        # day_count vars from group_assignments, NOT from profs.
        if a.group_id is not None:
            # Ensure the prof appears in `out` with an empty classi
            # so glibero / availability hooks still apply.
            out.setdefault(t.name, {"classi": {}, "glibero": []})
            continue
        cl = classes.get(support_class_id(a, students)
                         if a.is_support else a.class_id)
        if cl is None:
            continue
        node = out.setdefault(t.name, {"classi": {}, "glibero": []})
        slot = node["classi"].setdefault(cl.name, {})
        if a.is_support and a.subject in slot:
            # One teacher supporting two pupils of the same class is
            # two rows but ONE shadow cattedra: they share the solver
            # key (teacher, class, 'sostegno'), so the hours have to
            # add up instead of the second row silently replacing the
            # first. Same merge as in support_assignments_from_db.
            slot[a.subject]["ore"] += a.hours
        else:
            slot[a.subject] = {"ore": a.hours}
    # glibero + min_free_days HARD floor + per-day capacity
    day_caps = day_capacity_by_teacher(db)
    for tname, info in out.items():
        t = next((x for x in teachers.values() if x.name == tname), None)
        if t is None:
            continue
        primary = DAY_MAP.get(t.free_day or "Saturday", 6)
        rest = [d for d in days if d != primary]
        rng.shuffle(rest)
        info["glibero"] = [primary, rest[0], rest[1]]
        # Per-teacher >=N free-days floor. Solvers that don't preload
        # the DSL stream from a DB session (e.g. the benchmark harness)
        # read this off the profs dict so the stratified profile
        # distribution still reaches the model.
        info["min_free_days"] = int(
            getattr(t, "min_free_days", 1) or 1)
        # Per-day hour ceilings from the HARD availability tables.
        # Read off profs by the solvers that get no DB session --
        # notably cv2.solve_phase_a, whose day distribution is where
        # these ceilings actually bite. See day_capacity_by_teacher.
        info["day_capacity"] = {
            int(d): int(c)
            for d, c in sorted(day_caps.get(t.id, {}).items())
        }
    return out


# ============================================================
# Task C1: extensions for coteaching, sostegno, potenziamento
# ============================================================

def coteach_groups_for_solver(db: Session) -> list[dict]:
    """Return the list of CoteachGroup specs the solver consumes.

    Each entry:
      {
        'group_id': int,
        'class_name': str,
        'subject': str,
        'n_hours': int,
        'required': bool,
        'weight': float,
        'teachers': [principal_name, codoc_name1, codoc_name2, ...]
      }

    Members are derived from Assignment.coteach_group_id back-FK.
    Skipped silently if the group has fewer than 2 members.

    Convention enforced here: `teachers[0]` is the PRINCIPAL (the
    member with the most hours; ties broken alphabetically for
    determinism). Members[1:] are the CO-TEACHERS whose Assignment
    hours are exactly the n_hours of compresenza. The CP-SAT model
    in cpsat_v2_timetable.solve_phase_a relies on this ordering to
    apply day_count[principal] >= coday vs day_count[codoc] == coday.
    """
    out: list[dict] = []
    teachers_by_id = {t.id: t for t in db.query(models.Teacher).all()}
    classes_by_id = {c.id: c for c in db.query(models.SchoolClass).all()}
    studygroups_by_id = {
        sg.id: sg for sg in db.query(models.StudyGroup).all()
    }
    for g in db.query(models.CoteachGroup).all():
        members = [a for a in db.query(models.Assignment).filter(
            models.Assignment.coteach_group_id == g.id
        ).all() if a.teacher_id in teachers_by_id]
        if len(members) < 2:
            continue
        # Task C3: a CoteachGroup can target either a class
        # (g.class_id) or a StudyGroup (g.group_id). Resolve to the
        # solver-side label (class_name OR group_name).
        target_label: str | None = None
        if g.group_id is not None:
            sg = studygroups_by_id.get(g.group_id)
            if sg is None:
                continue
            target_label = sg.name
        elif g.class_id is not None:
            cl = classes_by_id.get(g.class_id)
            if cl is None:
                continue
            target_label = cl.name
        else:
            continue
        # Sort by hours DESC (principal first); ties broken by
        # teacher name for determinism.
        members_sorted = sorted(
            members,
            key=lambda a: (-int(a.hours or 0),
                            teachers_by_id[a.teacher_id].name),
        )
        teacher_names = [teachers_by_id[a.teacher_id].name
                         for a in members_sorted]
        out.append({
            "group_id": g.id,
            "class_name": target_label,
            "subject": g.subject,
            "n_hours": int(g.n_hours),
            "required": bool(g.required),
            "weight": float(g.weight or 100.0),
            "teachers": teacher_names,
        })
    return out


def support_assignments_from_db(db: Session) -> list[dict]:
    r"""Return the list of sostegno (shadow) assignments. Each entry:
      {
        'teacher_name': str, 'class_name': str,
        'subject': str (the SUPPORT_SUBJECT marker), 'n_hours': int,
        'is_group_target': bool,  # True if class_name is actually a
                                  # StudyGroup name (Task C3 sostegno
                                  # following a student in a group).
        'student_names': [str],   # the pupils actually followed
      }
    These rows have is_support=True and a target, resolved in this
    order: `student_id` (the pupil -> their current class, the normal
    shape), `group_id` (Task C3: sostegno follows the pupil into a
    StudyGroup), `class_id` (legacy class-level sostegno). The solver
    uses them to add
    `slot[sost,X,sost,h] <= OR(slot[*,X,*,h] for not-support)` for
    class-target, or `slot[sost,G,sost,h] <= group_busy[G,h]` for
    group-target, and to NOT count the sostegno slot in class-busy.

    Rows that land on the same (teacher, target, subject) are **merged
    with their hours summed** -- one teacher following two pupils of
    one class is two cattedre on paper but a single shadow in the
    model, since both would key the same solver variable.
    """
    merged: dict[tuple, dict] = {}
    teachers_by_id = {t.id: t for t in db.query(models.Teacher).all()}
    classes_by_id = {c.id: c for c in db.query(models.SchoolClass).all()}
    studygroups_by_id = {
        sg.id: sg for sg in db.query(models.StudyGroup).all()
    }
    students_by_id = {s.id: s for s in db.query(models.Student).all()}
    for a in db.query(models.Assignment).filter(
        models.Assignment.is_support == True  # noqa: E712
    ).all():
        t = teachers_by_id.get(a.teacher_id)
        if t is None:
            continue
        target_label: str | None = None
        is_group = False
        if a.group_id is not None:
            sg = studygroups_by_id.get(a.group_id)
            if sg is None:
                continue
            target_label = sg.name
            is_group = True
        else:
            cl = classes_by_id.get(support_class_id(a, students_by_id))
            if cl is None:
                continue
            target_label = cl.name
        st = students_by_id.get(getattr(a, "student_id", None))
        key = (t.name, target_label, a.subject, is_group)
        node = merged.get(key)
        if node is None:
            node = merged[key] = {
                "teacher_name": t.name,
                "class_name": target_label,
                "subject": a.subject,
                "n_hours": 0,
                "is_group_target": is_group,
                "student_names": [],
            }
        node["n_hours"] += int(a.hours)
        if st is not None:
            node["student_names"].append(
                f"{st.last_name} {st.first_name}".strip())
    return list(merged.values())


def parallel_groups_for_solver(db: Session) -> list[dict]:
    """Task C2: return the list of intra-class parallel groups
    (e.g. religione + alternativa). Each entry:
      {
        'group_id': int,
        'class_name': str,
        'members': [{teacher_name, subject, hours}, ...]
      }

    Built from Assignments grouped by (parallel_group_id, class_id).
    Skipped silently if a group has fewer than 2 members.
    """
    out: list[dict] = []
    teachers_by_id = {t.id: t for t in db.query(models.Teacher).all()}
    classes_by_id = {c.id: c for c in db.query(models.SchoolClass).all()}
    by_group: dict[tuple[int, int], list] = {}
    for a in db.query(models.Assignment).filter(
        models.Assignment.parallel_group_id != None  # noqa: E711
    ).all():
        if a.class_id is None:
            continue
        by_group.setdefault(
            (a.parallel_group_id, a.class_id), []).append(a)
    for (gid, cid), members in by_group.items():
        if len(members) < 2:
            continue
        cl = classes_by_id.get(cid)
        if cl is None:
            continue
        out.append({
            "group_id": gid,
            "class_name": cl.name,
            "members": [
                {"teacher_name": teachers_by_id[m.teacher_id].name,
                 "subject": m.subject,
                 "hours": int(m.hours or 0)}
                for m in members
                if m.teacher_id in teachers_by_id
            ],
        })
    return out


def potenziamento_assignments_from_db(db: Session) -> list[dict]:
    """Return the list of potenziamento (class-less) assignments.
      [{'teacher_name': str, 'subject': str, 'n_hours': int}]
    Skipped silently if class_id is set (data anomaly: a row marked
    is_potenziamento=True with a non-null class_id is malformed).
    """
    out: list[dict] = []
    teachers_by_id = {t.id: t for t in db.query(models.Teacher).all()}
    for a in db.query(models.Assignment).filter(
        models.Assignment.is_potenziamento == True  # noqa: E712
    ).all():
        if a.class_id is not None:
            continue                # malformed
        t = teachers_by_id.get(a.teacher_id)
        if t is None:
            continue
        out.append({
            "teacher_name": t.name,
            "subject": a.subject or "Potenziamento",
            "n_hours": int(a.hours),
        })
    return out


def group_assignments_for_solver(db: Session) -> list[dict]:
    """Task C3: return the list of inter-class group assignments.

    Each entry:
      {
        'teacher_name': str,
        'group_id': int,
        'group_name': str,        # used as virtual class label in solver
        'subject': str,
        'n_hours': int,
        'home_class_names': [str, ...]  # classes touched by group members
      }

    Built from Assignment rows with `group_id != NULL`. The solver
    treats the group as a virtual class with `group_slot[gid, d, h]`
    BoolVars; class_busy[home_class, d, h] is forced >= group_slot
    for every home class of any member student.

    `home_class_names` is the resolved list of (distinct) home classes
    of the students in the group. Empty list => the group has no
    members yet (the solver tolerates it but the assignment will not
    propagate class-busy anywhere).
    """
    out: list[dict] = []
    teachers_by_id = {t.id: t for t in db.query(models.Teacher).all()}
    students_by_id = {s.id: s for s in db.query(models.Student).all()}
    classes_by_id = {c.id: c for c in db.query(models.SchoolClass).all()}

    # Pre-compute home classes per group (one query, in-memory join).
    home_by_group: dict[int, list[str]] = {}
    for g in db.query(models.StudyGroup).all():
        home_class_ids: set[int] = set()
        for m in db.query(models.GroupMembership).filter(
            models.GroupMembership.group_id == g.id
        ).all():
            s = students_by_id.get(m.student_id)
            if s is None or s.class_id is None:
                continue
            home_class_ids.add(s.class_id)
        home_by_group[g.id] = sorted(
            cl.name for cid, cl in classes_by_id.items()
            if cid in home_class_ids
        )

    groups_by_id = {g.id: g for g in db.query(models.StudyGroup).all()}
    for a in db.query(models.Assignment).filter(
        models.Assignment.group_id != None  # noqa: E711
    ).all():
        t = teachers_by_id.get(a.teacher_id)
        g = groups_by_id.get(a.group_id)
        if t is None or g is None:
            continue
        out.append({
            "teacher_name": t.name,
            "group_id": g.id,
            "group_name": g.name,
            "subject": a.subject or "GroupSubject",
            "n_hours": int(a.hours or 0),
            "home_class_names": list(home_by_group.get(g.id, [])),
        })
    return out


def cattedre_from_assignments(db: Session) -> dict[str, dict[str, dict[str, int]]]:
    """teacher_name -> {class_name -> {subject -> ore}}"""
    out: dict[str, dict[str, dict[str, int]]] = {}
    teachers = {t.id: t.name for t in db.query(models.Teacher).all()}
    classes = {c.id: c.name for c in db.query(models.SchoolClass).all()}
    for a in db.query(models.Assignment).all():
        tn = teachers.get(a.teacher_id)
        cn = classes.get(a.class_id)
        if tn is None or cn is None:
            continue
        out.setdefault(tn, {}).setdefault(cn, {})[a.subject] = a.hours
    return out


# ---------- Pickles -> DB ----------


def import_school_into_db(db: Session, school: dict[str, Any],
                          replace: bool = True) -> None:
    """Replace classes + teachers + subject group weights with values from
    a school pickle dict (as produced by big_mock_school.build_dataset)."""
    if replace:
        for tbl in (
            models.Assignment,
            models.Lesson,
            models.DayCount,
            models.Solution,
            models.ClassSubject,
            models.SchoolClass,
            models.TeacherSubject,
            models.TeacherUnavailability,
            models.TeacherMandatoryFreeDay,
            models.TeacherCompatibleClass,
            models.Teacher,
            models.SubjectGroupWeight,
            models.Subject,
        ):
            db.query(tbl).delete()
        db.commit()

    # Subject rows
    subj_names: set[str] = set()
    for cl in school.get("classes", []):
        for s in cl.get("subjects", {}):
            subj_names.add(s)
    for t in school.get("teachers", []):
        for s in t.get("weights", {}):
            subj_names.add(s)
    for s in subj_names:
        db.add(models.Subject(name=s, pretty_name=s))
    db.flush()

    # Class rows
    for cl in school.get("classes", []):
        sc = models.SchoolClass(
            name=cl["name"],
            year=cl.get("year", 1),
            section=cl.get("section"),
            curriculum=cl.get("curriculum"),
            n_students=cl.get("n_students", 25),
        )
        db.add(sc)
        db.flush()
        for subj, ore in (cl.get("subjects") or {}).items():
            db.add(models.ClassSubject(
                class_id=sc.id, subject=subj, hours_per_week=int(ore)
            ))

    # Teacher rows
    for t in school.get("teachers", []):
        full = t["name"]
        # Best-effort split: Faker default emits "First Last" in en_US, while
        # Italian Faker emits "Last First" or "First Last" depending on the
        # locale's _format. Heuristic: rsplit on the LAST whitespace and treat
        # the right side as last_name; if there's only one token, leave them
        # as None.
        parts = str(full).rsplit(" ", 1)
        if len(parts) == 2:
            first_n, last_n = parts[0], parts[1]
        else:
            first_n, last_n = None, None
        tt = models.Teacher(
            name=full,
            last_name=last_n, first_name=first_n,
            nickname=None,
            group=t.get("group"),
            max_hours=int(t.get("max_hours", 18)),
            free_day=t.get("free_day"),
        )
        db.add(tt)
        db.flush()
        weights = t.get("weights") or {}
        for subj in weights:
            db.add(models.TeacherSubject(teacher_id=tt.id, subject=subj))

    # Subject group weight table
    for subj, gmap in (school.get("cconcorsopersubject") or {}).items():
        for g, w in gmap.items():
            db.add(models.SubjectGroupWeight(
                subject=subj, group_name=g, weight=int(w)
            ))
    db.commit()


def import_assignments_into_db(db: Session,
                               cattedre: dict[str, dict[str, dict[str, int]]]
                               ) -> int:
    """Replace existing assignments with the given mapping. Returns count."""
    db.query(models.Assignment).delete()
    db.commit()
    teachers = {t.name: t.id for t in db.query(models.Teacher).all()}
    classes = {c.name: c.id for c in db.query(models.SchoolClass).all()}
    n = 0
    for tname, cmap in cattedre.items():
        if tname not in teachers:
            # Auto-create teacher if missing (defensive: imported pickle
            # may carry teachers not in DB)
            t = models.Teacher(name=tname, max_hours=18)
            db.add(t)
            db.flush()
            teachers[tname] = t.id
        for cname, sm in cmap.items():
            if cname not in classes:
                continue
            for subj, ore in sm.items():
                db.add(models.Assignment(
                    teacher_id=teachers[tname],
                    class_id=classes[cname],
                    subject=subj,
                    hours=int(ore),
                ))
                n += 1
    db.commit()
    return n


def import_profs_into_db(db: Session, profs: dict[str, Any]) -> int:
    """Import a profs.pkl dict: each entry has classi -> class -> subject ->
    {'ore': N} and glibero. Updates teacher.free_day and assignments."""
    teachers = {t.name: t for t in db.query(models.Teacher).all()}
    classes = {c.name: c.id for c in db.query(models.SchoolClass).all()}
    db.query(models.Assignment).delete()
    db.commit()
    n = 0
    inv_day = {v: k for k, v in DAY_MAP.items()
               if not k.endswith("'") and "Lun" not in k}
    for tname, info in profs.items():
        t = teachers.get(tname)
        if t is None:
            t = models.Teacher(name=tname, max_hours=18)
            db.add(t)
            db.flush()
            teachers[tname] = t
        glib = info.get("glibero") or []
        if glib and not t.free_day:
            t.free_day = inv_day.get(glib[0], "Saturday")
        for cname, sm in (info.get("classi") or {}).items():
            cid = classes.get(cname)
            if cid is None:
                continue
            for subj, meta in sm.items():
                db.add(models.Assignment(
                    teacher_id=t.id,
                    class_id=cid,
                    subject=subj,
                    hours=int(meta.get("ore", 0)),
                ))
                n += 1
    db.commit()
    return n


# Italian-day-name map used by the orario_classi xlsx exporter. Kept
# narrow (only the abbreviations the exporter actually emits) so we do
# not accidentally match columns from a hand-edited xlsx with arbitrary
# headers.
_XLSX_DAY_TO_NUM = {
    "lun": 1, "lun.": 1, "lunedi": 1,
    "mar": 2, "mar.": 2, "martedi": 2,
    "mer": 3, "mer.": 3, "mercoledi": 3,
    "gio": 4, "gio.": 4, "giovedi": 4,
    "ven": 5, "ven.": 5, "venerdi": 5,
    "sab": 6, "sab.": 6, "sabato": 6,
}


def solution_dict_from_class_xlsx(xlsx_path: str) -> dict[tuple, int]:
    """Parse an ``orario_classi_<profile>.xlsx`` (as produced by
    ``engine.exporters.export_class_schedules_to_xlsx``) and rebuild the
    ``{(prof, class, subject, day, hour): 1}`` dict. Used as a last-resort
    fallback by the import-profile flow when the canonical solution pickle
    is gitignored / absent on disk but the xlsx is checked in.
    """
    try:
        from openpyxl import load_workbook
    except ImportError as e:  # pragma: no cover
        raise RuntimeError(f"openpyxl non disponibile: {e}") from e

    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    sol: dict[tuple, int] = {}
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        # A1: "Classe: <name>" — recover the original (un-truncated)
        # class name; fall back to the (possibly truncated) sheet name
        # when the title cell is missing.
        title = ws.cell(row=1, column=1).value or ""
        if isinstance(title, str) and title.lower().startswith("classe:"):
            class_name = title.split(":", 1)[1].strip()
        else:
            class_name = sheet_name

        # Header row 4: ['Ora', 'Lun', 'Mar', ...]; build column->day map.
        col_to_day: dict[int, int] = {}
        for c in range(2, 16):
            h = ws.cell(row=4, column=c).value
            if h is None:
                continue
            key = str(h).strip().lower().rstrip(".")
            d = _XLSX_DAY_TO_NUM.get(key)
            if d is not None:
                col_to_day[c] = d
        if not col_to_day:
            continue  # not a schedule sheet; skip

        # Hour rows: column A holds "1^a (08:00)"; we use the leading
        # integer as the hour. Stop on the first blank A-cell.
        for r in range(5, 5 + 12):
            hour_cell = ws.cell(row=r, column=1).value
            if hour_cell is None or str(hour_cell).strip() == "":
                break
            digits = ""
            for ch in str(hour_cell):
                if ch.isdigit():
                    digits += ch
                else:
                    if digits:
                        break
            if not digits:
                continue
            hour = int(digits)
            for col, day in col_to_day.items():
                cell = ws.cell(row=r, column=col).value
                if cell is None:
                    continue
                txt = str(cell).strip()
                if not txt or txt == "-":
                    continue
                if txt.startswith("*CONFLICT*"):
                    # The exporter writes "*CONFLICT* subj1/prof1 ; subj2/prof2"
                    # when multiple lessons collided on a slot. Reconstructing
                    # both halves would create duplicate Lesson rows for the
                    # same (class, day, hour); skip and let the user re-run
                    # the optimizer for a clean schedule.
                    continue
                if "/" not in txt:
                    continue
                subj, prof = txt.split("/", 1)
                subj = subj.strip()
                prof = prof.strip()
                if not subj or not prof:
                    continue
                sol[(prof, class_name, subj, day, hour)] = 1
    return sol


def import_solution_into_db(db: Session, solution_dict: dict,
                            name: str, kind: str = "imported",
                            obj_value: float = 0.0,
                            metrics: dict | None = None,
                            make_active: bool = True) -> int:
    """Persist a (p,cl,subj,day,hour) -> 0/1 dict as a Solution row + many
    Lesson rows. Returns the new solution id."""
    sol = models.Solution(
        name=name,
        kind=kind,
        obj_value=float(obj_value),
        metrics_json=json.dumps(metrics or {}),
        is_active=False,
        created_at=dt.datetime.utcnow(),
    )
    db.add(sol)
    db.flush()
    for k, v in solution_dict.items():
        if v != 1:
            continue
        try:
            p, cl, subj, day, hour = k
        except Exception:
            continue
        db.add(models.Lesson(
            solution_id=sol.id,
            teacher_name=p,
            class_name=cl,
            subject=subj,
            day=int(day),
            hour=int(hour),
        ))
    db.commit()
    if make_active:
        set_active_solution(db, sol.id)
    return sol.id


def set_active_solution(db: Session, solution_id: int) -> None:
    db.query(models.Solution).update({"is_active": False})
    sol = db.get(models.Solution, solution_id)
    if sol is not None:
        sol.is_active = True
    db.commit()


def synthesize_solution_from_profs(profs: dict[str, Any]) -> dict[tuple, int]:
    """Greedy round-robin schedule: distribute each cattedra (teacher,
    class, subject, hours) across the week respecting only teacher and
    class non-overlap. Output is a solution_dict shaped exactly like
    what `import_solution_into_db` consumes:

        {(teacher, class, subject, day, hour): 1, ...}

    Used as a fallback when an engine profile (school + profs) is
    imported but no solver-produced ``solution_*.pkl`` is on disk yet.
    The placeholder schedule is HARD-feasible w.r.t. teacher/class
    overlap (ignores room and SOFT objectives), so /schedule renders a
    full grid the user can then optimise via Phase B.
    """
    DAYS, HOURS = list(range(1, 7)), list(range(8, 14))
    slots = [(d, h) for d in DAYS for h in HOURS]
    sol: dict[tuple, int] = {}
    teacher_busy: dict[str, set[tuple[int, int]]] = defaultdict(set)
    class_busy: dict[str, set[tuple[int, int]]] = defaultdict(set)

    cattedre: list[tuple[int, str, str, str, set[int]]] = []
    for tname, info in (profs or {}).items():
        glib = {int(d) for d in (info.get("glibero") or []) if str(d).isdigit() or isinstance(d, int)}
        for cname, sm in (info.get("classi") or {}).items():
            for subj, meta in (sm or {}).items():
                ore = int((meta or {}).get("ore", 0))
                if ore > 0:
                    cattedre.append((ore, str(tname), str(cname),
                                     str(subj), glib))
    cattedre.sort(reverse=True)

    for ore, tname, cname, subj, glib in cattedre:
        placed = 0
        for (d, h) in slots:
            if placed >= ore:
                break
            if d in glib:
                continue
            if (d, h) in teacher_busy[tname]:
                continue
            if (d, h) in class_busy[cname]:
                continue
            sol[(tname, cname, subj, d, h)] = 1
            teacher_busy[tname].add((d, h))
            class_busy[cname].add((d, h))
            placed += 1
    return sol


def ensure_default_working_hours(db: Session, tenant_id: int = 1) -> int:
    """Defensive: re-seed the canonical lun-sab + 8:00-14:00
    WorkingDay/WorkingHourSlot configuration if the tenant has none.
    Returns the number of inserted day rows (0 if already populated).

    init_db() seeds these on a fresh schema; this helper is called
    from the engine-profile importer in case a previous wipe or a
    botched migration left the working-hours tables empty -- without
    them /schedule's grid layout ends up empty even though Lessons
    exist."""
    existing = db.query(models.WorkingDay).filter(
        models.WorkingDay.tenant_id == tenant_id
    ).count()
    if existing:
        return 0
    defaults = [
        ("MON", "Lunedi",    0, 1),
        ("TUE", "Martedi",   1, 2),
        ("WED", "Mercoledi", 2, 3),
        ("THU", "Giovedi",   3, 4),
        ("FRI", "Venerdi",   4, 5),
        ("SAT", "Sabato",    5, 6),
    ]
    for code, label, pos, legacy in defaults:
        d = models.WorkingDay(
            tenant_id=tenant_id, code=code, label=label,
            position=pos, legacy_day_number=legacy, is_active=True,
        )
        db.add(d)
        db.flush()
        for i in range(6):
            h = 8 + i
            db.add(models.WorkingHourSlot(
                day_id=d.id, slot_index=i,
                start_time=f"{h:02d}:00", end_time=f"{h+1:02d}:00",
                label=f"{i+1}ª ora", legacy_hour_number=h,
            ))
    db.commit()
    return len(defaults)


def get_active_solution(db: Session) -> models.Solution | None:
    return db.query(models.Solution).filter(
        models.Solution.is_active == True  # noqa: E712
    ).first()


def lessons_to_solution_dict(db: Session, solution_id: int
                             ) -> dict[tuple, int]:
    out: dict[tuple, int] = {}
    rows = db.query(models.Lesson).filter(
        models.Lesson.solution_id == solution_id
    ).all()
    for r in rows:
        out[(r.teacher_name, r.class_name, r.subject, r.day, r.hour)] = 1
    return out


def replace_solution_lessons(db: Session, solution_id: int,
                             solution_dict: dict[tuple, int]) -> None:
    """Carry over classroom_name/co-teachers from previous Lesson rows when
    keys still match — this preserves manual room edits across small SOFT
    repairs (the keys are stable in (p, cl, subj, d, h))."""
    prev = db.query(models.Lesson).filter(
        models.Lesson.solution_id == solution_id
    ).all()
    prev_by_key = {
        (l.teacher_name, l.class_name, l.subject, l.day, l.hour):
        (l.classroom_name, l.cotaught_with) for l in prev
    }
    db.query(models.Lesson).filter(
        models.Lesson.solution_id == solution_id
    ).delete()
    db.commit()
    for k, v in solution_dict.items():
        if v != 1:
            continue
        p, cl, subj, day, hour = k
        room, cot = prev_by_key.get((p, cl, subj, day, hour),
                                    (None, None))
        db.add(models.Lesson(
            solution_id=solution_id,
            teacher_name=p,
            class_name=cl,
            subject=subj,
            day=int(day),
            hour=int(hour),
            classroom_name=room,
            cotaught_with=cot,
        ))
    db.commit()


# ---------- Classrooms ----------


def class_flags_from_db(db: Session) -> dict[str, dict[str, bool]]:
    """Per-class overrides of the seven ex-officio HARD invariants
    (finding 08b): ``{class_name -> {no_holes, entry_at_8, exit_after_12,
    dual_math, dual_italian, motorie_pairs, max_6_per_day}}``.

    The SchoolClass columns default True (the historical global
    behaviour), so a class the school never touched keeps every invariant;
    only a class where a toggle was explicitly turned OFF differs. The
    solver reads this and gates the matching per-class constraint, instead
    of applying all seven to every class regardless of the class card.
    """
    out: dict[str, dict[str, bool]] = {}
    for c in db.query(models.SchoolClass).all():
        out[c.name] = {
            "no_holes": bool(getattr(c, "hard_no_holes", True)),
            "entry_at_8": bool(getattr(c, "hard_entry_at_8", True)),
            "exit_after_12": bool(getattr(c, "hard_exit_after_12", True)),
            "dual_math": bool(getattr(c, "hard_dual_math", True)),
            "dual_italian": bool(getattr(c, "hard_dual_italian", True)),
            "motorie_pairs": bool(getattr(c, "hard_motorie_pairs", True)),
            "max_6_per_day": bool(getattr(c, "hard_max_6_per_day", True)),
        }
    return out


def classrooms_dicts_from_db(db: Session) -> list[dict]:
    """Materialize classrooms with all their constraints, ready to feed
    classroom_assignment.solve_classroom_assignment()."""
    out: list[dict] = []
    for r in db.query(models.Classroom).all():
        unav = {(u.day, u.hour) for u in r.unavailability}
        subj_req = {sp.subject for sp in r.subject_prefs if sp.required}
        subj_forbidden = {sp.subject for sp in r.subject_prefs
                          if sp.state == "forbidden"}
        # `state` e\` la fonte di verita\`, il SEGNO di `weight` no: la UI
        # scrive -20 per 'preferred' (convenzione "costo negativo"), il
        # generatore mock scrive +10. Qui usciamo con una sola
        # convenzione -- magnitudine POSITIVA = quanto e\` gradita --
        # cosi\` il consumatore non deve indovinare. Solo le righe
        # 'preferred' entrano: 'enforced' e\` gia\` HARD in
        # `subject_required`, 'forbidden' e\` un divieto (non una
        # preferenza negativa) e 'allowed' e\` neutro.
        subj_pref = {sp.subject: abs(float(sp.weight or 0.0))
                     for sp in r.subject_prefs
                     if sp.state == "preferred" and sp.weight}
        cls_pref = {cp.class_name: abs(float(cp.weight or 0.0))
                    for cp in r.class_prefs
                    if cp.state == "preferred" and cp.weight}
        is_home_for = {cp.class_name for cp in r.class_prefs if cp.is_home}
        out.append({
            "name": r.name,
            "kind": r.kind,
            "capacity": r.capacity,
            "multi_class": r.multi_class,
            "multi_class_max": r.multi_class_max,
            "multi_class_pref": r.multi_class_pref,
            "multi_class_pref_weight": r.multi_class_pref_weight,
            "unavailability": unav,
            "subject_required": subj_req,
            "subject_forbidden": subj_forbidden,
            "subject_pref_weight": subj_pref,
            "class_pref_weight": cls_pref,
            "is_home_for": is_home_for,
        })
    return out


def room_pins_from_db(db: Session) -> dict:
    r"""Risolve i preset di ``SchoolClass.room_policy`` e gli stati HARD
    di ``ClassroomClassPreference`` in due mappe per classe.

    Ritorna ``{"pin": {classe: aula}, "forbidden": {classe: {aule}},
    "fissa_senza_aula": [classi]}``.

    - ``pin`` e\` l'aula in cui la classe deve stare (HARD). Vale per
      ``room_policy='fissa'`` (l'aula base) e per ogni riga esplicita
      con ``state='enforced'``, che ha la precedenza: e\` il caso
      manuale "questa classe sta li\`, non nella sua aula base".
    - ``forbidden`` sono le aule vietate alla classe.
    - ``fissa_senza_aula`` elenca le classi con preset 'fissa' ma senza
      riga ``is_home``: non c'e\` niente da fissare, quindi si degrada a
      'ibrida'. Il chiamante lo segnala nel log invece di far fallire
      il run -- un preset senza il dato che gli serve e\` un errore di
      configurazione, non un motivo per non produrre l'orario.
    """
    policy_by_class = {
        c.name: (c.room_policy or "ibrida")
        for c in db.query(models.SchoolClass).all()
    }
    room_name_by_id = {
        r.id: r.name for r in db.query(models.Classroom).all()
    }
    home_by_class: dict[str, str] = {}
    pin: dict[str, str] = {}
    forbidden: dict[str, set] = defaultdict(set)
    for cp in db.query(models.ClassroomClassPreference).all():
        room = room_name_by_id.get(cp.classroom_id)
        if not room:
            continue
        state = cp.state or "preferred"
        if state == "forbidden":
            forbidden[cp.class_name].add(room)
        elif state == "enforced":
            pin[cp.class_name] = room
        if cp.is_home:
            home_by_class.setdefault(cp.class_name, room)

    fissa_senza_aula = []
    for cl, policy in policy_by_class.items():
        if policy != "fissa" or cl in pin:
            continue
        home = home_by_class.get(cl)
        if home:
            pin[cl] = home
        else:
            fissa_senza_aula.append(cl)
    return {
        "pin": pin,
        "forbidden": {k: v for k, v in forbidden.items()},
        "fissa_senza_aula": sorted(fissa_senza_aula),
    }


def compresenza_resolver(db: Session):
    r"""Ritorna un predicato ``(teacher_name, day, hour) -> bool``: quel
    docente, in quella cella, sta in compresenza?

    Legge solo ``Teacher.compresenza`` (+ ``TeacherCompresenzaHour`` per
    il modo 'oraria'). NON guarda ``Assignment.is_support``: il sostegno
    e\` gia\` stato tradotto in ``compresenza='sempre'`` a monte, quando
    la cattedra e\` stata creata (o dalla migrazione a4d81e2c9f57 per i
    DB preesistenti). Cosi\` la stessa configurazione copre sostegno,
    potenziamento, codocenza, madrelingua e ITP senza casi speciali.

    Attenzione: 'sempre' e\` una proprieta\` del DOCENTE, non della
    singola lezione. Un docente di sostegno che abbia anche una
    cattedra ordinaria risponderebbe True anche li\`. Il chiamante deve
    quindi trattare la risposta come "puo\` accodarsi a una lezione
    ospite", non come "non occupa un'aula": senza un ospite nella
    stessa classe e ora, la lezione torna a prenotare un'aula propria
    (vedi ``classroom_assignment.solve_classroom_assignment``).
    """
    teachers = db.query(models.Teacher).all()
    name_by_id = {t.id: t.name for t in teachers}
    modes = {t.name: (getattr(t, "compresenza", None) or "mai")
             for t in teachers}
    cells: dict[str, set[tuple[int, int]]] = {}
    if any(m == "oraria" for m in modes.values()):
        for r in db.query(models.TeacherCompresenzaHour).all():
            nm = name_by_id.get(r.teacher_id)
            if nm is None:
                continue
            cells.setdefault(nm, set()).add((int(r.day), int(r.hour)))

    def _shares(teacher_name: str, day, hour) -> bool:
        mode = modes.get(teacher_name, "mai")
        if mode == "sempre":
            return True
        if mode == "oraria":
            return (int(day), int(hour)) in cells.get(teacher_name, ())
        return False

    return _shares


_ROOM_CONTINUITY_PRAGMAS = (
    ("class_same_room_per_day", "day"),
    ("class_same_room_per_week", "week"),
    ("class_room_changes_min", "soft"),
)


def room_continuity_from_dsl(db: Session) -> dict:
    r"""Parse room-continuity pragmas out of the stored general-DSL
    constraints and return ``{class_name -> "day" | "week" | "soft"}``.

    The pragmas are the single mechanism the user asked for through TWO
    doors: a saved ``GeneralConstraint`` row is a *preset*, and the same
    text typed into the generic-DSL box is the ad-hoc form. Recognised:

    * ``class_same_room_per_day(<class>)``  -> HARD same room within a day
    * ``class_same_room_per_week(<class>)`` -> HARD same room all week
    * ``class_room_changes_min(<class>)``   -> SOFT minimise room changes

    A class named by more than one pragma keeps the strictest (day > week
    > soft). Room vars only exist in the joint model, so these are applied
    there (see ``classroom_assignment.add_room_continuity_constraints``);
    unlike ordinary general-DSL rules they are NOT evaluated post-hoc.
    """
    try:
        rows = db.query(models.GeneralConstraint).all()
    except Exception:
        return {}
    return parse_room_continuity_pragmas(
        getattr(r, "expression", "") or "" for r in rows)


def parse_room_continuity_pragmas(expressions) -> dict:
    r"""Pure parser (no DB): scan DSL ``expressions`` for the three
    room-continuity pragmas and return ``{class -> mode}``, strictest
    wins (day > week > soft)."""
    import re
    rank = {"day": 3, "week": 2, "soft": 1}
    out: dict[str, str] = {}
    for expr in expressions:
        for name, mode in _ROOM_CONTINUITY_PRAGMAS:
            for m in re.finditer(name + r"\(\s*([^)]+?)\s*\)", expr or ""):
                cl = m.group(1).strip().strip("'\"").strip()
                if not cl:
                    continue
                if cl not in out or rank[mode] > rank[out[cl]]:
                    out[cl] = mode
    return out


def joint_room_ctx_from_db(db: Session) -> dict:
    r"""Placement-INDEPENDENT room metadata for the joint (day,hour,room)
    solver -- everything ``classroom_assignment.add_joint_room_vars``
    needs that does NOT depend on where lessons land.

    Unlike :func:`lessons_for_classroom_step` (which reads a materialized
    solution), this is built before Phase B has decided the timetable, so
    the joint model can constrain the schedule to be room-feasible while
    it is still choosing hours. Returns a dict with the classroom dicts,
    ``required_kind`` per subject, the HARD ``home_room`` / ``forbidden``
    per class (from the ``room_policy`` preset), and the compresenza
    ``shares`` predicate used to drop rider cells.
    """
    pins = room_pins_from_db(db)
    return {
        "classrooms": classrooms_dicts_from_db(db),
        "required_kind_by_subj": {
            s.name: (s.required_kind or "")
            for s in db.query(models.Subject).all()
        },
        "home_by_class": dict(pins["pin"]),
        "forbidden_by_class": dict(pins["forbidden"]),
        "shares": compresenza_resolver(db),
        "continuity_dsl": room_continuity_from_dsl(db),
    }


def joint_cells_from_slot_keys(slot_keys, ctx: dict):
    r"""Turn the schedule model's ``slot`` keys ``(teacher, class,
    subject, day, hour)`` into the two maps the joint room primitive
    consumes: ``cell_to_keys`` ``{(cl,subj,d,h) -> [slot_keys]}`` and
    ``cell_lessons`` ``{(cl,subj,d,h) -> lesson_meta}``.

    Compresenza RIDER cells are dropped (they inherit the host's room
    downstream, exactly as the standalone solver's ``compresenza_map``
    does). A cell is a rider only when *all* its teachers share a room
    AND there is a non-rider host cell in the same ``(class, day,
    hour)`` -- without a host the lesson books its own room, matching
    ``compresenza_resolver``'s own caveat.
    """
    from collections import defaultdict
    shares = ctx["shares"]
    req_kind = ctx["required_kind_by_subj"]
    home = ctx["home_by_class"]
    forbidden = ctx["forbidden_by_class"]

    cell_to_keys: dict[tuple, list] = defaultdict(list)
    cell_teachers: dict[tuple, list] = defaultdict(list)
    for key in slot_keys:
        t, cl, s, d, h = key
        cell = (cl, s, int(d), int(h))
        cell_to_keys[cell].append(key)
        cell_teachers[cell].append(t)

    # Candidate riders: every teacher on the cell is in compresenza there.
    candidate_rider = {
        cell for cell, ts in cell_teachers.items()
        if ts and all(shares(t, cell[2], cell[3]) for t in ts)
    }
    # Promote to TRUE rider only when a host shares the (class, day, hour).
    riders: set = set()
    by_class_slot: dict[tuple, list] = defaultdict(list)
    for cell in cell_teachers:
        by_class_slot[(cell[0], cell[2], cell[3])].append(cell)
    for cells in by_class_slot.values():
        if any(c not in candidate_rider for c in cells):
            riders.update(c for c in cells if c in candidate_rider)

    cell_lessons: dict[tuple, dict] = {}
    for cell, ts in cell_teachers.items():
        if cell in riders:
            continue
        cl, s, d, h = cell
        cell_lessons[cell] = {
            "class": cl, "subject": s, "day": d, "hour": h,
            "required_kind": req_kind.get(s, "") or "",
            "home_room": home.get(cl, "") or "",
            "forbidden_rooms": forbidden.get(cl, set()),
            "teacher": ts[0] if ts else "",
            "co_teachers": list(ts[1:]),
        }
    out_keys = {c: ks for c, ks in cell_to_keys.items() if c not in riders}
    return out_keys, cell_lessons


def lessons_for_classroom_step(db: Session, solution_id: int,
                               *, pins: dict | None = None) -> list[dict]:
    r"""Convert lessons in the active solution to the shape expected by
    classroom_assignment.

    `n_students` per lesson is looked up from SchoolClass and used by
    classroom_assignment to enforce the HARD capacity constraint
    (room.capacity >= class.n_students). Classes missing from the DB
    (defensive: should not happen on a clean import) get a 0 so the
    constraint is trivially satisfied.

    `pins` e\` il risultato di `room_pins_from_db`; viene calcolato qui
    se il chiamante non lo passa gia\` pronto.
    """
    n_students_by_class = {
        c.name: int(c.n_students or 0)
        for c in db.query(models.SchoolClass).all()
    }
    required_kind_by_subj = {
        s.name: (s.required_kind or None)
        for s in db.query(models.Subject).all()
    }
    if pins is None:
        pins = room_pins_from_db(db)
    pin_by_class = pins.get("pin", {})
    forbidden_by_class = pins.get("forbidden", {})
    shares = compresenza_resolver(db)
    out = []
    for l in db.query(models.Lesson).filter(
        models.Lesson.solution_id == solution_id
    ).all():
        co = []
        if l.cotaught_with:
            co = [s.strip() for s in l.cotaught_with.split(",") if s.strip()]
        out.append({
            "teacher": l.teacher_name,
            "co_teachers": co,
            "class": l.class_name,
            "subject": l.subject,
            "day": l.day,
            "hour": l.hour,
            "n_students": n_students_by_class.get(l.class_name, 0),
            # Pulled from Subject.required_kind. NULL/missing -> ''
            # which classroom_assignment treats as "any room".
            "required_kind": required_kind_by_subj.get(l.subject, "") or "",
            # Aula base HARD (preset 'fissa' o riga 'enforced'). ''
            # significa nessun vincolo. La deroga per le materie con
            # required_kind e\` applicata dentro `_can_host`, non qui:
            # cosi\` il dato che esce dal DB resta la configurazione
            # dichiarata e non una sua interpretazione.
            "home_room": pin_by_class.get(l.class_name, "") or "",
            "forbidden_rooms": forbidden_by_class.get(l.class_name, set()),
            # Compresenza: la lezione si accoda all'aula di un'altra
            # lezione della stessa classe e ora invece di prenotarne
            # una propria. Vedi `compresenza_resolver`.
            "shares_room": bool(shares(l.teacher_name, l.day, l.hour)),
        })
    return out


def apply_room_mapping(db: Session, solution_id: int,
                       mapping: dict[tuple, str]) -> int:
    """Updates Lesson.classroom_name from a (class, subject, day, hour)
    -> classroom_name dict. Returns the number of rows updated."""
    n = 0
    for l in db.query(models.Lesson).filter(
        models.Lesson.solution_id == solution_id
    ).all():
        key = (l.class_name, l.subject, l.day, l.hour)
        if key in mapping:
            l.classroom_name = mapping[key]
            n += 1
    db.commit()
    return n


# ---------- SQLite profile snapshot import ----------


# Tables copied verbatim from the profile SQLite into the live DB.
# Order matters: parent rows before children. Lessons + Solution come
# last because they reference school_classes (by name) implicitly.
_PROFILE_TABLES = [
    "subjects",
    "subject_group_weights",
    "teachers",
    "school_classes",
    "class_subjects",
    "teacher_subjects",
    "teacher_unavailability",
    "teacher_mandatory_free_days",
    "teacher_compatible_classes",
    "teacher_class_preferences",
    "teacher_curriculum_preferences",
    "curricula",
    "curriculum_subject_hours",
    "curriculum_logical_constraints",
    "plessi",
    "plesso_commuting_rules",
    "plesso_entity_policies",
    "classrooms",
    "classroom_subject_preferences",
    "classroom_class_preferences",
    "classroom_unavailability",
    "classroom_tags",
    "classroom_tag_assignments",
    "class_unavailability",
    "logical_unavailabilities",
    "coteach_groups",
    "assignments",
    "general_constraints",
    "working_days",
    "working_hour_slots",
    "solutions",
    "lessons",
]


# Tables wiped in REVERSE order before the import to avoid FK issues.
_PROFILE_TABLES_WIPE = list(reversed(_PROFILE_TABLES)) + [
    # Solutions / Run / DayCount aren't part of the profile snapshot
    # but they're meaningful at import time -- the live DB shouldn't
    # carry stale runs after a fresh profile import. Lessons get
    # wiped above; we add Run + DayCount + UnscheduledLesson.
    "day_counts",
    "unscheduled_lessons",
]


def import_profile_sqlite_into_db(
    db: Session, sqlite_path: str, *, replace: bool = True,
    replace_working_hours: bool = True,
) -> dict[str, int]:
    """Copy every relevant table from a profile SQLite snapshot into
    the live DB.

    Reads the source SQLite via ``sqlite3``; writes through the live
    Session's raw DBAPI connection (``exec_driver_sql`` accepts ``?``
    placeholders so positional tuple binds work with SQLAlchemy 2.x).

    The schemas were created from the same ``models.py`` (the build
    script uses ``Base.metadata.create_all``), so a column-by-column
    copy is safe; we project to the intersection of columns present
    in both DBs to stay resilient to migrations that ran on the live
    DB but not on the snapshot (or vice versa).

    Returns ``{table: rows_inserted}`` for the run log.
    """
    import sqlite3
    if not os.path.exists(sqlite_path):
        raise FileNotFoundError(sqlite_path)

    src = sqlite3.connect(sqlite_path)
    src.row_factory = sqlite3.Row
    counts: dict[str, int] = {}
    try:
        from sqlalchemy import text as _text
        from sqlalchemy import inspect as sa_inspect
        live_insp = sa_inspect(db.get_bind())
        live_tables = set(live_insp.get_table_names())

        if replace:
            for t in _PROFILE_TABLES_WIPE:
                if not replace_working_hours and t in (
                        "working_hour_slots", "working_days"):
                    continue
                if t in live_tables:
                    db.execute(_text(f"DELETE FROM {t}"))
            db.commit()

        # Use the underlying DBAPI connection for ``?``-style binds.
        raw = db.connection().connection
        for t in _PROFILE_TABLES:
            if not replace_working_hours and t in (
                    "working_hour_slots", "working_days"):
                continue
            if t not in live_tables:
                continue
            live_cols = {c["name"] for c in live_insp.get_columns(t)}
            src_cols = [r[1] for r in src.execute(
                f"PRAGMA table_info({t})").fetchall()]
            if not src_cols:
                continue
            shared = [c for c in src_cols if c in live_cols]
            if not shared:
                continue
            # Quote identifiers -- some columns (e.g. teachers.group)
            # collide with SQL reserved words.
            quoted = [f'"{c}"' for c in shared]
            rows = src.execute(
                f"SELECT {','.join(quoted)} FROM {t}").fetchall()
            if not rows:
                continue
            placeholders = ",".join(["?"] * len(shared))
            cols_q = ",".join(quoted)
            stmt = f"INSERT INTO {t} ({cols_q}) VALUES ({placeholders})"
            cur = raw.cursor()
            try:
                cur.executemany(
                    stmt, [tuple(row[c] for c in shared)
                           for row in rows])
                counts[t] = len(rows)
            finally:
                cur.close()
        db.commit()
    finally:
        src.close()
    return counts
