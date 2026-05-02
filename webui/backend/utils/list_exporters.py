"""Generic list-to-xlsx / list-to-csv exporters.

Used by every list endpoint that exposes a `?format=xlsx` or
`?format=csv` query parameter (teachers, classes, classrooms,
subjects, students, groups, ...). The list endpoints already
applied the user's `q=` and `sort=` filters; the exporter just
streams the resulting list of dicts as a tabular file.
"""
from __future__ import annotations

import csv
import io
from datetime import datetime
from typing import Any, Iterable, Sequence


# ---------- Column resolution -----------------------------------------

def _flatten_value(v: Any) -> Any:
    """Coerce non-primitive values (lists, dicts) into a string so the
    cell stays tabular. Empty -> ''."""
    if v is None:
        return ""
    if isinstance(v, bool):
        return "1" if v else "0"
    if isinstance(v, (int, float)):
        return v
    if isinstance(v, str):
        return v
    if isinstance(v, list):
        # Try to stringify list of primitives nicely; else json-ish repr
        try:
            return ", ".join(
                str(x) if not isinstance(x, dict) else _flatten_value(x)
                for x in v
            )
        except Exception:
            return str(v)
    if isinstance(v, dict):
        # Display dicts as "k=v;k=v"
        try:
            return "; ".join(f"{k}={_flatten_value(val)}"
                              for k, val in v.items())
        except Exception:
            return str(v)
    return str(v)


def _infer_columns(rows: Sequence[dict]) -> list[str]:
    """Stable union of keys preserving first-seen order."""
    seen: dict[str, None] = {}
    for r in rows:
        for k in r.keys():
            if k not in seen:
                seen[k] = None
    return list(seen.keys())


def _resolve_columns(rows: Sequence[dict],
                     columns: Sequence[str] | None) -> list[str]:
    if columns is not None:
        return list(columns)
    return _infer_columns(rows)


# ---------- Public exports --------------------------------------------

def to_xlsx(rows: Sequence[dict],
            entity: str,
            columns: Sequence[str] | None = None,
            sheet_title: str | None = None) -> bytes:
    """Returns the raw bytes of an xlsx workbook with one sheet
    listing the rows in the given column order. If `columns` is None,
    the union of keys (insertion order) is used.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    cols = _resolve_columns(rows, columns)
    wb = Workbook()
    ws = wb.active
    if ws is not None:
        ws.title = (sheet_title or entity)[:31]
        # Header
        ws.append(cols)
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill("solid", fgColor="2563EB")
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        # Data
        for r in rows:
            ws.append([_flatten_value(r.get(c)) for c in cols])
        # Naive column-width auto-fit (capped)
        for i, c in enumerate(cols, start=1):
            longest = max(
                [len(str(c))]
                + [len(str(_flatten_value(r.get(c))))
                   for r in rows[:200]],
                default=10,
            )
            ws.column_dimensions[
                ws.cell(row=1, column=i).column_letter
            ].width = min(60, max(8, longest + 2))

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def to_csv(rows: Sequence[dict],
           entity: str,
           columns: Sequence[str] | None = None) -> bytes:
    """Returns CSV bytes (UTF-8 with BOM for Excel compatibility on
    Italian locale). Comma separator. Quoting per RFC 4180.
    """
    cols = _resolve_columns(rows, columns)
    buf = io.StringIO()
    writer = csv.writer(buf, delimiter=",", quoting=csv.QUOTE_MINIMAL,
                         lineterminator="\r\n")
    writer.writerow(cols)
    for r in rows:
        writer.writerow([_flatten_value(r.get(c)) for c in cols])
    out = buf.getvalue().encode("utf-8")
    return b"\xef\xbb\xbf" + out  # BOM


def export_filename(entity: str, fmt: str) -> str:
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    return f"{entity}_{ts}.{fmt}"


XLSX_MIME = ("application/"
             "vnd.openxmlformats-officedocument.spreadsheetml.sheet")
CSV_MIME = "text/csv; charset=utf-8"


def respond_export(rows: Iterable[dict], entity: str,
                   fmt: str,
                   columns: Sequence[str] | None = None):
    """Build a FastAPI Response (StreamingResponse) for a list export.

    Caller is responsible for having already applied q/sort filters
    on `rows`. `entity` is used in the filename.
    """
    from fastapi.responses import Response
    rows = list(rows)
    fmt = (fmt or "").lower()
    if fmt == "xlsx":
        body = to_xlsx(rows, entity, columns)
        media_type = XLSX_MIME
    elif fmt == "csv":
        body = to_csv(rows, entity, columns)
        media_type = CSV_MIME
    else:
        raise ValueError(f"format non supportato: {fmt}")
    fn = export_filename(entity, fmt)
    return Response(
        content=body,
        media_type=media_type,
        headers={"Content-Disposition": f'attachment; filename="{fn}"'},
    )
