r"""Modulo riusabile di export dell'orario settimanale in Excel.

API
===

    export_class_schedules_to_xlsx(
        solution_pickle_path: str,
        school_pickle_path:   str,
        profs_pickle_path:    str,
        output_path:          str,
    ) -> None

    export_teacher_schedules_to_xlsx(
        solution_pickle_path: str,
        school_pickle_path:   str,
        profs_pickle_path:    str,
        output_path:          str,
    ) -> None

Entrambe leggono:
- `solution_pickle_path`: dict[(prof, classe, materia, day, hour) -> 0/1]
- `school_pickle_path`  : dict prodotto da `big_mock_school.py`
                          (campi: classes, teachers, ...). Usato per
                          arricchire gli header (curriculum della classe,
                          classe-di-concorso del docente, max_hours).
- `profs_pickle_path`   : dict prof -> {classi: {cl: {subj: {ore}}},
                          glibero: [d1,d2,d3]}.

Producono un singolo `.xlsx` in `output_path`. Non rilanciano alcun
solver: sono solo trasformazione di dati.

Edge case gestiti:
- nomi tab Excel > 31 char: troncati a 28 + "..."; collisioni di nome
  dopo troncamento: suffisso numerico "~1", "~2", ... mantenendo la
  lunghezza max 31.
- caratteri Excel proibiti `[]:*?/\\` sostituiti con "-".
- docente con piu\` materie nella stessa classe: nel tab del docente
  la cella mostra "<classe>\n(<materia>)" se ambiguo.
- docente con piu\` materie totali: header "Materie:" elenca tutte.
- compresenza nello stesso slot (non dovrebbe mai accadere visto il
  vincolo hard no-overlap, ma e\` difensivo): cella prefissata
  "*CONFLICT*" e l'evento viene loggato (numero totale a fine).

Utilizzabili anche da CLI -- vedi il blocco `if __name__ ...` in fondo.
"""
from __future__ import annotations

import os
import pickle
import re

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


# --------------------------- costanti di stile ---------------------------

DAY_NAMES: dict[int, str] = {
    1: "Lun", 2: "Mar", 3: "Mer", 4: "Gio", 5: "Ven", 6: "Sab",
}
EMPTY_PLACEHOLDER = "-"

_HEADER_FILL = PatternFill(
    start_color="4472C4", end_color="4472C4", fill_type="solid"
)
_HEADER_FONT = Font(bold=True, color="FFFFFF")
_TITLE_FONT = Font(bold=True, size=14)
_SUBTITLE_FONT = Font(italic=True, color="555555")
_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
_LEFT = Alignment(horizontal="left", vertical="center")
_THIN = Side(border_style="thin", color="999999")
_CELL_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_EMPTY_FILL = PatternFill(
    start_color="F2F2F2", end_color="F2F2F2", fill_type="solid"
)
_CONFLICT_FILL = PatternFill(
    start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"
)

_FORBIDDEN_TAB_CHARS = re.compile(r"[\[\]:\*\?/\\]")


# --------------------------- helpers ---------------------------

def _safe_sheet_name(name: str, used: set[str]) -> str:
    """Ritorna un nome di tab Excel valido (<=31 char, no chars vietati)
    e non duplicato."""
    cleaned = _FORBIDDEN_TAB_CHARS.sub("-", name).strip() or "Sheet"
    if len(cleaned) > 31:
        base = cleaned[:28] + "..."
    else:
        base = cleaned
    candidate = base
    i = 1
    while candidate in used:
        suffix = f"~{i}"
        keep = 31 - len(suffix)
        candidate = (base[:keep] if len(base) > keep else base) + suffix
        i += 1
    used.add(candidate)
    return candidate


def _load_pickles(solution_path, school_path, profs_path):
    with open(solution_path, "rb") as f:
        sol = pickle.load(f)
    with open(school_path, "rb") as f:
        school = pickle.load(f)
    with open(profs_path, "rb") as f:
        profs = pickle.load(f)
    return sol, school, profs


def _classes_index(school) -> dict[str, dict]:
    """Indice nome_classe -> dict del classroom dump."""
    return {c["name"]: c for c in school.get("classes", [])}


def _teachers_index(school) -> dict[str, dict]:
    """Indice nome_docente -> dict del teacher dump (group, max_hours...)."""
    return {t["name"]: t for t in school.get("teachers", [])}


def _days_hours(sol) -> tuple[list[int], list[int]]:
    days = sorted({k[3] for k in sol})
    hours = sorted({k[4] for k in sol})
    return days, hours


def _write_grid_header(ws, days, header_row: int):
    cell = ws.cell(row=header_row, column=1, value="Ora")
    cell.fill = _HEADER_FILL
    cell.font = _HEADER_FONT
    cell.alignment = _CENTER
    cell.border = _CELL_BORDER
    for i, d in enumerate(days):
        c = ws.cell(row=header_row, column=2 + i,
                    value=DAY_NAMES.get(d, f"Day{d}"))
        c.fill = _HEADER_FILL
        c.font = _HEADER_FONT
        c.alignment = _CENTER
        c.border = _CELL_BORDER


def _write_hour_column(ws, hours, start_row: int):
    for i, h in enumerate(hours):
        ord_label = f"{i + 1}^a ({h:02d}:00)"
        c = ws.cell(row=start_row + i, column=1, value=ord_label)
        c.fill = _HEADER_FILL
        c.font = _HEADER_FONT
        c.alignment = _CENTER
        c.border = _CELL_BORDER


def _fill_cell(ws, row, col, text: str, conflict: bool = False):
    c = ws.cell(row=row, column=col, value=text)
    c.alignment = _CENTER
    c.border = _CELL_BORDER
    if conflict:
        c.fill = _CONFLICT_FILL
        c.font = Font(bold=True, color="9C0006")
    elif text == EMPTY_PLACEHOLDER:
        c.fill = _EMPTY_FILL


def _autosize(ws, days, *, header_row: int, hour_row_count: int,
              col_a_width: float = 14, col_other_width: float = 26,
              header_h: float = 26, body_h: float = 36):
    ws.column_dimensions["A"].width = col_a_width
    for i in range(len(days)):
        ws.column_dimensions[get_column_letter(2 + i)].width = col_other_width
    for r in range(1, header_row + 1):
        ws.row_dimensions[r].height = header_h
    for r in range(header_row + 1, header_row + 1 + hour_row_count):
        ws.row_dimensions[r].height = body_h


# --------------------------- funzioni pubbliche ---------------------------

def export_class_schedules_to_xlsx(
    solution_pickle_path: str,
    school_pickle_path: str,
    profs_pickle_path: str,
    output_path: str,
) -> None:
    """Crea un .xlsx con un tab per classe.

    Colonne = giorni della settimana (Lun..Sab in base al mock).
    Righe = ore della giornata (1^a, 2^a, ..., con orario assoluto).
    Celle = "Materia / Docente" o "-" se la classe non ha lezione.
    """
    sol, school, profs = _load_pickles(
        solution_pickle_path, school_pickle_path, profs_pickle_path
    )
    cls_idx = _classes_index(school)
    days, hours = _days_hours(sol)
    classes = sorted({c for p in profs.values() for c in p["classi"]})

    wb = Workbook()
    wb.remove(wb.active)
    used = set()
    n_conflicts = 0

    for cl in classes:
        sheet_name = _safe_sheet_name(cl, used)
        ws = wb.create_sheet(title=sheet_name)
        cls_meta = cls_idx.get(cl, {})
        curriculum = cls_meta.get("curriculum", "")

        # Header informativo
        ws.cell(row=1, column=1, value=f"Classe: {cl}").font = _TITLE_FONT
        ws.merge_cells(start_row=1, start_column=1,
                       end_row=1, end_column=1 + len(days))
        if curriculum:
            ws.cell(row=2, column=1,
                    value=f"Indirizzo: {curriculum}").font = _SUBTITLE_FONT
            ws.merge_cells(start_row=2, start_column=1,
                           end_row=2, end_column=1 + len(days))
        header_row = 4

        _write_grid_header(ws, days, header_row=header_row)
        _write_hour_column(ws, hours, start_row=header_row + 1)

        for di, day in enumerate(days):
            for hi, hour in enumerate(hours):
                lessons = []
                for prof, info in profs.items():
                    if cl not in info["classi"]:
                        continue
                    for subj in info["classi"][cl]:
                        if sol.get((prof, cl, subj, day, hour), 0) == 1:
                            lessons.append((subj, prof))
                row = header_row + 1 + hi
                col = 2 + di
                if not lessons:
                    _fill_cell(ws, row, col, EMPTY_PLACEHOLDER)
                elif len(lessons) == 1:
                    subj, prof = lessons[0]
                    _fill_cell(ws, row, col, f"{subj} / {prof}")
                else:
                    n_conflicts += 1
                    text = "*CONFLICT* " + " ; ".join(
                        f"{s}/{p}" for s, p in lessons
                    )
                    _fill_cell(ws, row, col, text, conflict=True)

        _autosize(ws, days, header_row=header_row,
                  hour_row_count=len(hours))
        # freeze: blocca l'header e la colonna ore
        ws.freeze_panes = ws.cell(row=header_row + 1, column=2).coordinate

    out_dir = os.path.dirname(os.path.abspath(output_path))
    if out_dir:
        os.makedirs(out_dir, exist_ok=True)
    wb.save(output_path)
    print(
        f"[exporters] {output_path}: "
        f"{len(wb.sheetnames)} tab classi, "
        f"{n_conflicts} conflitti rilevati"
    )


def export_teacher_schedules_to_xlsx(
    solution_pickle_path: str,
    school_pickle_path: str,
    profs_pickle_path: str,
    output_path: str,
) -> None:
    r"""Crea un .xlsx con un tab per docente.

    Header del tab (righe 1-3): nome docente, materie insegnate, totali.
    Colonne = giorni; righe = ore; celle = nome della classe.
    Se il docente insegna piu\` materie nella stessa classe, la cella
    riporta "<classe>\n(<materia>)" per disambiguare.
    """
    sol, school, profs = _load_pickles(
        solution_pickle_path, school_pickle_path, profs_pickle_path
    )
    teachers_idx = _teachers_index(school)
    days, hours = _days_hours(sol)

    wb = Workbook()
    wb.remove(wb.active)
    used = set()
    n_conflicts = 0

    for prof in sorted(profs.keys()):
        sheet_name = _safe_sheet_name(prof, used)
        ws = wb.create_sheet(title=sheet_name)
        info = profs[prof]
        meta = teachers_idx.get(prof, {})
        group = meta.get("group", "")
        max_hours = meta.get("max_hours", "?")

        subj_set = sorted({
            s for cl, sm in info["classi"].items() for s in sm
        })
        n_classes = len(info["classi"])
        tot_ore = sum(
            v["ore"] for cl, sm in info["classi"].items()
            for v in sm.values()
        )

        # Header info
        title = f"Docente: {prof}"
        if group:
            title += f"   (cl. concorso {group})"
        ws.cell(row=1, column=1, value=title).font = _TITLE_FONT
        ws.merge_cells(start_row=1, start_column=1,
                       end_row=1, end_column=1 + len(days))
        ws.cell(row=2, column=1,
                value=f"Materie insegnate: "
                      f"{', '.join(subj_set) if subj_set else '(nessuna)'}"
                ).font = _SUBTITLE_FONT
        ws.merge_cells(start_row=2, start_column=1,
                       end_row=2, end_column=1 + len(days))
        ws.cell(
            row=3, column=1,
            value=f"Classi: {n_classes} | "
                  f"Ore assegnate: {tot_ore} | "
                  f"Disponibilita\\` max: {max_hours}"
        ).font = _SUBTITLE_FONT
        ws.merge_cells(start_row=3, start_column=1,
                       end_row=3, end_column=1 + len(days))
        header_row = 5

        _write_grid_header(ws, days, header_row=header_row)
        _write_hour_column(ws, hours, start_row=header_row + 1)

        for di, day in enumerate(days):
            for hi, hour in enumerate(hours):
                lessons = []
                for cl in info["classi"]:
                    for subj in info["classi"][cl]:
                        if sol.get((prof, cl, subj, day, hour), 0) == 1:
                            lessons.append((cl, subj))
                row = header_row + 1 + hi
                col = 2 + di
                if not lessons:
                    _fill_cell(ws, row, col, EMPTY_PLACEHOLDER)
                elif len(lessons) == 1:
                    cl, subj = lessons[0]
                    if len(info["classi"][cl]) > 1:
                        text = f"{cl}\n({subj})"
                    else:
                        text = cl
                    _fill_cell(ws, row, col, text)
                else:
                    n_conflicts += 1
                    text = "*CONFLICT* " + " ; ".join(
                        f"{c}/{s}" for c, s in lessons
                    )
                    _fill_cell(ws, row, col, text, conflict=True)

        _autosize(
            ws, days,
            header_row=header_row, hour_row_count=len(hours),
            col_other_width=22,
        )
        ws.freeze_panes = ws.cell(row=header_row + 1, column=2).coordinate

    out_dir = os.path.dirname(os.path.abspath(output_path))
    if out_dir:
        os.makedirs(out_dir, exist_ok=True)
    wb.save(output_path)
    print(
        f"[exporters] {output_path}: "
        f"{len(wb.sheetnames)} tab docenti, "
        f"{n_conflicts} conflitti rilevati"
    )


# --------------------------- esempio CLI ---------------------------

if __name__ == "__main__":
    here = os.path.dirname(os.path.abspath(__file__))
    profile = "huge"
    out_dir = os.path.join(here, "output", profile)

    sol_path = os.path.join(here, f"solution_timetable_{profile}.pkl")
    school_path = os.path.join(here, f"school_{profile}.pkl")
    profs_path = os.path.join(here, f"profs_{profile}.pkl")

    export_class_schedules_to_xlsx(
        sol_path, school_path, profs_path,
        os.path.join(out_dir, f"orario_classi_{profile}.xlsx"),
    )
    export_teacher_schedules_to_xlsx(
        sol_path, school_path, profs_path,
        os.path.join(out_dir, f"orario_docenti_{profile}.xlsx"),
    )
